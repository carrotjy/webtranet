from flask import Blueprint, request, jsonify, send_file
import os
import shutil
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, Color

# PDF 변환을 위한 임포트
import subprocess
import platform

# WeasyPrint를 사용한 순수 Python PDF 생성 (Linux/Mac에서만 사용)
# Windows에서는 GTK 의존성 문제로 사용 불가
HAS_WEASYPRINT = False
try:
    # Windows가 아닌 경우에만 import 시도
    if platform.system() != 'Windows':
        from weasyprint import HTML
        from jinja2 import Template, Environment, FileSystemLoader
        HAS_WEASYPRINT = True
        print("WeasyPrint 사용 가능 (Linux/Mac)")
    else:
        print("Windows 환경: WeasyPrint 비활성화 (GTK 의존성 문제)")
        print("프로덕션(Ubuntu)에서는 WeasyPrint가 자동으로 활성화됩니다.")
except ImportError as e:
    HAS_WEASYPRINT = False
    print(f"WeasyPrint를 사용할 수 없습니다: {str(e)}")
except Exception as e:
    HAS_WEASYPRINT = False
    print(f"WeasyPrint 로드 실패: {str(e)}")

invoice_generator_bp = Blueprint('invoice_generator', __name__)

# 상수 정의
INSTANCE_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'instance')
INVOICE_BASE_DIR = os.path.join(INSTANCE_DIR, '거래명세서')
TEMPLATE_PATH = os.path.join(INSTANCE_DIR, '거래명세표(cs양식).xlsx')

def ensure_customer_folder(customer_name: str) -> str:
    """고객사 폴더 생성 또는 확인 - instance/거래명세서/고객사명"""
    # 거래명세서 폴더 생성
    if not os.path.exists(INVOICE_BASE_DIR):
        os.makedirs(INVOICE_BASE_DIR)

    # 고객사 폴더 생성
    customer_folder = os.path.join(INVOICE_BASE_DIR, customer_name)
    if not os.path.exists(customer_folder):
        os.makedirs(customer_folder)
    return customer_folder

def get_or_create_invoice_file(customer_name: str) -> str:
    """거래명세표 파일 경로 반환 (파일 존재 여부 확인만)"""
    customer_folder = ensure_customer_folder(customer_name)
    invoice_file_path = os.path.join(customer_folder, f'거래명세표({customer_name}).xlsx')
    return invoice_file_path

def copy_template_file(customer_name: str) -> tuple:
    """
    템플릿 파일 전체를 복사하여 고객사별 파일 생성 (서식 100% 보존)
    Returns: (workbook, invoice_file_path)
    """
    import shutil

    invoice_file_path = get_or_create_invoice_file(customer_name)

    # 파일이 없으면 템플릿 전체 복사
    if not os.path.exists(invoice_file_path):
        shutil.copy2(TEMPLATE_PATH, invoice_file_path)
        print(f"템플릿 파일 복사: {TEMPLATE_PATH} -> {invoice_file_path}")

    # 복사된 파일 열기
    workbook = load_workbook(invoice_file_path)

    return workbook, invoice_file_path

def prepare_sheet_from_template(workbook, sheet_name: str) -> tuple:
    """
    워크북 내의 Template-customer와 Template-supplier 시트를 복사하여 새 시트 생성
    ★ workbook.copy_worksheet() 사용으로 서식 100% 보존 ★

    Returns:
        tuple: (customer_sheet, supplier_sheet)
    """
    template_customer_name = 'Template-customer'
    template_supplier_name = 'Template-supplier'

    # 템플릿 시트 확인
    if template_customer_name not in workbook.sheetnames:
        raise Exception(f'워크북에 {template_customer_name} 시트를 찾을 수 없습니다.')
    if template_supplier_name not in workbook.sheetnames:
        raise Exception(f'워크북에 {template_supplier_name} 시트를 찾을 수 없습니다.')

    # Template 시트 가져오기
    template_customer_sheet = workbook[template_customer_name]
    template_supplier_sheet = workbook[template_supplier_name]

    # Template 시트들 숨김 처리
    template_customer_sheet.sheet_state = 'hidden'
    template_supplier_sheet.sheet_state = 'hidden'
    print(f"{template_customer_name}, {template_supplier_name} 시트를 숨김 처리했습니다.")

    # 동일한 이름의 시트가 있으면 -1, -2 등을 추가
    def get_unique_sheet_name(base_name):
        final_name = base_name
        counter = 1
        while final_name in workbook.sheetnames:
            final_name = f"{base_name}-{counter}"
            counter += 1
        return final_name

    # 고객사용 시트 생성 (yymmdd-c) - workbook.copy_worksheet() 사용으로 서식 100% 보존
    customer_sheet_name = get_unique_sheet_name(f"{sheet_name}-c")
    customer_sheet = workbook.copy_worksheet(template_customer_sheet)
    customer_sheet.title = customer_sheet_name
    print(f"고객사용 시트 생성: {customer_sheet_name}")

    # 공급자용 시트 생성 (yymmdd-s) - workbook.copy_worksheet() 사용으로 서식 100% 보존
    supplier_sheet_name = get_unique_sheet_name(f"{sheet_name}-s")
    supplier_sheet = workbook.copy_worksheet(template_supplier_sheet)
    supplier_sheet.title = supplier_sheet_name
    print(f"공급자용 시트 생성: {supplier_sheet_name}")

    # 새 시트들을 맨 앞으로 이동 (customer가 먼저, supplier가 그 다음)
    workbook.move_sheet(supplier_sheet, offset=-len(workbook.sheetnames) + 1)
    workbook.move_sheet(customer_sheet, offset=-len(workbook.sheetnames) + 1)

    return (customer_sheet, supplier_sheet)

def copy_template_sheet_to_target_OLD(target_workbook, sheet_name: str) -> object:
    """
    [구버전 - 사용 안 함]
    템플릿 파일에서 Template 시트를 복사하여 타겟 워크북에 붙여넣기
    문제: 인쇄 설정 등이 완벽하게 복사되지 않음
    """
    # 템플릿 파일 열기
    template_workbook = load_workbook(TEMPLATE_PATH)
    template_sheet_name = 'Template'

    # Template 시트가 없으면 에러
    if template_sheet_name not in template_workbook.sheetnames:
        template_workbook.close()
        raise Exception(f'템플릿 파일에 {template_sheet_name} 시트를 찾을 수 없습니다.')

    # Template 시트 가져오기
    template_sheet = template_workbook[template_sheet_name]

    # 동일한 이름의 시트가 있으면 -1, -2 등을 추가
    final_sheet_name = sheet_name
    counter = 1
    while final_sheet_name in target_workbook.sheetnames:
        final_sheet_name = f"{sheet_name}-{counter}"
        counter += 1

    # 타겟 워크북에 새 시트 생성
    new_sheet = target_workbook.create_sheet(final_sheet_name)

    # 템플릿 시트의 사용된 영역만 복사 (서식 포함)
    if template_sheet.max_row > 0:
        for row in template_sheet.iter_rows(min_row=1, max_row=template_sheet.max_row,
                                            min_col=1, max_col=template_sheet.max_column):
            for cell in row:
                new_cell = new_sheet[cell.coordinate]
                # 값 복사
                if cell.value is not None:
                    new_cell.value = cell.value
                # 서식 복사 (값이 없어도 서식은 복사)
                if cell.has_style:
                    new_cell.font = cell.font.copy()
                    new_cell.border = cell.border.copy()
                    new_cell.fill = cell.fill.copy()
                    new_cell.number_format = cell.number_format
                    new_cell.protection = cell.protection.copy()
                    new_cell.alignment = cell.alignment.copy()

    # 열 너비 복사
    for col_letter in template_sheet.column_dimensions:
        if col_letter in template_sheet.column_dimensions:
            new_sheet.column_dimensions[col_letter].width = template_sheet.column_dimensions[col_letter].width

    # 행 높이 복사
    for row_num in template_sheet.row_dimensions:
        if row_num in template_sheet.row_dimensions:
            new_sheet.row_dimensions[row_num].height = template_sheet.row_dimensions[row_num].height

    # 병합된 셀 복사
    for merged_cell_range in template_sheet.merged_cells.ranges:
        new_sheet.merge_cells(str(merged_cell_range))

    # 페이지 설정 복사 (인쇄 영역, 여백, 용지 크기 등)
    if template_sheet.page_setup:
        new_sheet.page_setup.orientation = template_sheet.page_setup.orientation
        new_sheet.page_setup.paperSize = template_sheet.page_setup.paperSize
        new_sheet.page_setup.fitToPage = template_sheet.page_setup.fitToPage
        new_sheet.page_setup.fitToHeight = template_sheet.page_setup.fitToHeight
        new_sheet.page_setup.fitToWidth = template_sheet.page_setup.fitToWidth
        new_sheet.page_setup.scale = template_sheet.page_setup.scale

    # 인쇄 여백 복사
    if template_sheet.page_margins:
        new_sheet.page_margins.left = template_sheet.page_margins.left
        new_sheet.page_margins.right = template_sheet.page_margins.right
        new_sheet.page_margins.top = template_sheet.page_margins.top
        new_sheet.page_margins.bottom = template_sheet.page_margins.bottom
        new_sheet.page_margins.header = template_sheet.page_margins.header
        new_sheet.page_margins.footer = template_sheet.page_margins.footer

    # 인쇄 옵션 복사
    if template_sheet.print_options:
        new_sheet.print_options.horizontalCentered = template_sheet.print_options.horizontalCentered
        new_sheet.print_options.verticalCentered = template_sheet.print_options.verticalCentered
        new_sheet.print_options.headings = template_sheet.print_options.headings
        new_sheet.print_options.gridLines = template_sheet.print_options.gridLines

    # 시트 보기 설정 복사 (확대/축소, 고정 틀 등)
    if template_sheet.sheet_view and template_sheet.sheet_view.sheetViewList:
        new_sheet.sheet_view = template_sheet.sheet_view

    # 인쇄 영역 복사
    if hasattr(template_sheet, 'print_area') and template_sheet.print_area:
        try:
            # 인쇄 영역은 워크북 레벨에서 정의됨
            if template_sheet.print_area:
                # 기존 시트 이름을 새 시트 이름으로 변경
                print_area_str = str(template_sheet.print_area)
                print_area_str = print_area_str.replace(f"'{template_sheet.title}'!", f"'{final_sheet_name}'!")
                print_area_str = print_area_str.replace(f"{template_sheet.title}!", f"{final_sheet_name}!")

                # 워크북에 인쇄 영역 정의 추가
                target_workbook.defined_names.append(
                    type('DefinedName', (), {
                        'name': '_xlnm.Print_Area',
                        'localSheetId': target_workbook.sheetnames.index(final_sheet_name),
                        'value': print_area_str
                    })()
                )
        except Exception as e:
            print(f"인쇄 영역 복사 실패: {str(e)}")

    # 템플릿 워크북 닫기
    template_workbook.close()

    # 새 시트를 맨 앞으로 이동
    target_workbook.move_sheet(new_sheet, offset=-len(target_workbook.sheetnames) + 1)

    return new_sheet

def write_header_info_to_sheet(sheet, service_date: str, customer_info: dict, total_amount: int):
    """헤더 정보를 시트에 기입 (작성일자, 고객사명, 주소, 전화/팩스, 합계금액)"""
    # 병합된 셀에 값을 쓰기 전에 병합 해제 필요 여부 확인
    def safe_write_to_cell(cell_ref, value):
        """병합된 셀을 안전하게 처리하며 값 입력"""
        try:
            # 병합된 셀 범위 확인
            for merged_range in sheet.merged_cells.ranges:
                if cell_ref in merged_range:
                    # 병합 범위의 왼쪽 상단 셀 좌표 가져오기
                    min_col, min_row, max_col, max_row = merged_range.bounds
                    top_left_cell = sheet.cell(row=min_row, column=min_col)
                    top_left_cell.value = value
                    return
            # 병합되지 않은 셀이면 직접 입력
            sheet[cell_ref] = value
        except Exception as e:
            print(f"셀 {cell_ref}에 값 입력 실패: {str(e)}")
            # 병합 해제 후 재시도
            try:
                sheet.unmerge_cells(cell_ref)
                sheet[cell_ref] = value
            except:
                pass

    # B4: 작성일자 (yyyy-mm-dd 형식 - 양식에 의해 자동 포맷팅됨)
    if service_date:
        safe_write_to_cell('B4', service_date)

    # F5: 고객사명
    if customer_info.get('company_name'):
        safe_write_to_cell('F5', customer_info.get('company_name'))

    # F7: 주소
    if customer_info.get('address'):
        safe_write_to_cell('F7', customer_info.get('address'))

    # F9: 전화번호와 팩스번호 (두 줄 형태)
    phone_fax = []
    if customer_info.get('phone'):
        phone_fax.append(customer_info.get('phone'))
    if customer_info.get('fax'):
        phone_fax.append(customer_info.get('fax'))
    if phone_fax:
        safe_write_to_cell('F9', '\n'.join(phone_fax))

    # F11: 합계금액 (0이 아닌 경우에만)
    if total_amount and total_amount != 0:
        safe_write_to_cell('F11', total_amount)

def apply_thick_border_to_range(sheet, start_cell: str = 'B3', end_cell: str = 'AG43'):
    """
    지정된 영역에 두꺼운 테두리 적용
    주변 테두리 색상과 동일하게 설정
    """
    from openpyxl.utils import range_boundaries

    # 기존 테두리 색상 가져오기 (B3 셀의 테두리 색상 사용)
    sample_cell = sheet['B3']
    border_color = Color(rgb='000000')  # 기본값은 검정색

    # 샘플 셀에서 테두리 색상 객체를 직접 사용
    if sample_cell.border and sample_cell.border.left and sample_cell.border.left.color:
        border_color = sample_cell.border.left.color

    # 두꺼운 테두리 스타일 정의 (medium)
    thick_side = Side(style='medium', color=border_color)

    # 범위 파싱
    min_col, min_row, max_col, max_row = range_boundaries(f"{start_cell}:{end_cell}")

    # 범위의 모든 셀에 테두리 적용
    for row_idx in range(min_row, max_row + 1):
        for col_idx in range(min_col, max_col + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)

            # 기존 테두리 복사
            if cell.border:
                left = cell.border.left
                right = cell.border.right
                top = cell.border.top
                bottom = cell.border.bottom
            else:
                left = right = top = bottom = Side(style=None)

            # 테두리 영역의 가장자리인 경우 두꺼운 테두리 적용
            # 왼쪽 가장자리
            if col_idx == min_col:
                left = thick_side
            # 오른쪽 가장자리
            if col_idx == max_col:
                right = thick_side
            # 위쪽 가장자리
            if row_idx == min_row:
                top = thick_side
            # 아래쪽 가장자리
            if row_idx == max_row:
                bottom = thick_side

            # 새 테두리 적용
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

def write_invoice_items_to_sheet(sheet, items: list, start_row: int = 14):
    """거래명세서 항목을 시트에 기입"""
    current_row = start_row

    # 병합된 셀 안전 처리 함수
    def safe_write_to_cell(cell_ref, value):
        """병합된 셀을 안전하게 처리하며 값 입력"""
        try:
            # 병합된 셀 범위 확인
            for merged_range in sheet.merged_cells.ranges:
                if cell_ref in merged_range:
                    # 병합 범위의 왼쪽 상단 셀 좌표 가져오기
                    min_col, min_row, max_col, max_row = merged_range.bounds
                    top_left_cell = sheet.cell(row=min_row, column=min_col)
                    top_left_cell.value = value
                    return top_left_cell
            # 병합되지 않은 셀이면 직접 입력
            sheet[cell_ref] = value
            return sheet[cell_ref]
        except Exception as e:
            print(f"셀 {cell_ref}에 값 입력 실패: {str(e)}")
            return None

    for item in items:
        # 빈 행은 건너뛰기 (행만 차지)
        if item.get('isBlank'):
            current_row += 1
            continue

        # 헤더 행은 월, 일, 품목만 입력
        if item.get('isHeader'):
            if item.get('month'):
                safe_write_to_cell(f'B{current_row}', item.get('month'))
            if item.get('day'):
                safe_write_to_cell(f'C{current_row}', item.get('day'))
            safe_write_to_cell(f'D{current_row}', item.get('item_name'))
            current_row += 1
            continue

        # 일반 데이터 행
        # 열 주소: B=월, C=일, D=품목, J=규격, O=수량, Q=단가, V=공급가액, AA=부가세

        # 월 (0이 아닌 경우에만)
        month = item.get('month')
        if month and month != 0:
            safe_write_to_cell(f'B{current_row}', month)

        # 일 (0이 아닌 경우에만)
        day = item.get('day')
        if day and day != 0:
            safe_write_to_cell(f'C{current_row}', day)

        # 품목 (필수)
        if item.get('item_name'):
            safe_write_to_cell(f'D{current_row}', item.get('item_name'))

        # 규격 (값이 있을 때만)
        if item.get('specification'):
            safe_write_to_cell(f'J{current_row}', item.get('specification'))

        # 수량 (0이 아닌 경우에만)
        quantity = item.get('quantity')
        if quantity and quantity != 0:
            safe_write_to_cell(f'O{current_row}', quantity)

        # 단가 (0이 아닌 경우에만)
        unit_price = item.get('unit_price')
        if unit_price and unit_price != 0:
            safe_write_to_cell(f'Q{current_row}', unit_price)

        # NEGO 항목은 마이너스로 저장
        total_price = item.get('total_price', 0)
        vat = item.get('vat', 0)
        if item.get('item_name') == 'NEGO':
            total_price = -abs(total_price) if total_price else 0
            vat = -abs(vat) if vat else 0

        # 공급가액 (0이 아닌 경우에만)
        if total_price and total_price != 0:
            safe_write_to_cell(f'V{current_row}', total_price)

        # 부가세 (0이 아닌 경우에만)
        if vat and vat != 0:
            safe_write_to_cell(f'AA{current_row}', vat)

        # NEGO 항목인 경우 빨간색 폰트 적용 (기존 폰트 속성 유지)
        if item.get('item_name') == 'NEGO':
            for col in ['B', 'C', 'D', 'J', 'O', 'Q', 'V', 'AA']:
                cell_ref = f'{col}{current_row}'
                try:
                    # 병합된 셀 확인
                    target_cell = None
                    for merged_range in sheet.merged_cells.ranges:
                        if cell_ref in merged_range:
                            min_col, min_row, max_col, max_row = merged_range.bounds
                            target_cell = sheet.cell(row=min_row, column=min_col)
                            break
                    else:
                        # 병합되지 않은 셀
                        target_cell = sheet[cell_ref]

                    if target_cell and target_cell.font:
                        # 기존 폰트 속성을 복사하고 색상만 빨간색으로 변경
                        original_font = target_cell.font
                        target_cell.font = Font(
                            name=original_font.name,
                            size=original_font.size,
                            bold=original_font.bold,
                            italic=original_font.italic,
                            vertAlign=original_font.vertAlign,
                            underline=original_font.underline,
                            strike=original_font.strike,
                            color="FF0000"  # 빨간색만 변경
                        )
                except:
                    pass

        current_row += 1

    return current_row

def calculate_total_amount(items: list) -> int:
    """합계금액 계산 (NEGO 항목은 마이너스로)"""
    total_price = 0
    total_vat = 0

    for item in items:
        if item.get('item_name') == 'NEGO':
            total_price -= item.get('total_price', 0)
            total_vat -= item.get('vat', 0)
        else:
            total_price += item.get('total_price', 0)
            total_vat += item.get('vat', 0)

    return total_price + total_vat

def convert_to_pdf_weasyprint(items: list, service_date: str, customer_info: dict, total_amount: int, pdf_path: str) -> bool:
    """
    WeasyPrint를 사용하여 HTML에서 PDF 생성 (Linux/Mac 전용)
    Windows에서는 GTK 의존성 문제로 사용 불가
    """
    if not HAS_WEASYPRINT:
        print("WeasyPrint를 사용할 수 없습니다. (Windows 또는 설치되지 않음)")
        return False

    try:
        from weasyprint import HTML
        from jinja2 import Environment, FileSystemLoader

        # Jinja2 템플릿 로드
        template_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'templates')
        env = Environment(loader=FileSystemLoader(template_dir))
        template = env.get_template('invoice_template.html')

        # 합계 계산
        total_supply = 0
        total_vat = 0
        for item in items:
            if item.get('isHeader') or item.get('isBlank'):
                continue
            if item.get('item_name') == 'NEGO':
                total_supply -= item.get('total_price', 0)
                total_vat -= item.get('vat', 0)
            else:
                total_supply += item.get('total_price', 0)
                total_vat += item.get('vat', 0)

        # 템플릿 렌더링
        html_content = template.render(
            service_date=service_date,
            customer_info=customer_info,
            items=items,
            total_amount=total_amount,
            total_supply=total_supply,
            total_vat=total_vat,
            generation_time=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        )

        # PDF 생성
        HTML(string=html_content).write_pdf(pdf_path)

        print(f"WeasyPrint PDF 생성 성공: {pdf_path}")
        return True

    except Exception as e:
        print(f"WeasyPrint PDF 변환 실패: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def convert_excel_to_pdf_libreoffice(excel_path: str, pdf_path: str, sheet_name: str = None) -> bool:
    """
    LibreOffice를 사용하여 Excel 파일을 PDF로 변환
    우분투/리눅스에서 작동하며 Windows에서도 LibreOffice가 설치되어 있으면 작동

    Args:
        excel_path: 원본 Excel 파일 경로
        pdf_path: 생성할 PDF 파일 경로
        sheet_name: 특정 시트만 PDF로 변환 (None이면 전체)
    """
    try:
        output_dir = os.path.dirname(os.path.abspath(pdf_path))

        # 특정 시트만 PDF로 변환하는 경우, 임시 파일 생성
        if sheet_name:
            import tempfile
            workbook = load_workbook(excel_path)

            if sheet_name not in workbook.sheetnames:
                print(f"경고: 시트 '{sheet_name}'를 찾을 수 없습니다.")
                workbook.close()
                return False

            # 임시 파일에 해당 시트만 복사
            temp_wb = load_workbook(excel_path)  # 전체 파일 복사

            # 타겟 시트를 제외한 모든 시트 삭제
            sheets_to_remove = [s for s in temp_wb.sheetnames if s != sheet_name]
            for sheet in sheets_to_remove:
                del temp_wb[sheet]

            # 임시 파일로 저장
            temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
            temp_excel_path = temp_file.name
            temp_file.close()

            temp_wb.save(temp_excel_path)
            temp_wb.close()
            workbook.close()

            abs_excel_path = os.path.abspath(temp_excel_path)
            print(f"임시 파일 생성: {temp_excel_path} (시트: {sheet_name}만 포함)")
        else:
            abs_excel_path = os.path.abspath(excel_path)
            temp_excel_path = None

        # 운영체제에 따라 LibreOffice 명령어 찾기
        system = platform.system()

        if system == 'Windows':
            # Windows에서 LibreOffice 경로 시도
            possible_paths = [
                r'C:\Program Files\LibreOffice\program\soffice.exe',
                r'C:\Program Files (x86)\LibreOffice\program\soffice.exe',
                'soffice.exe'  # PATH에 있는 경우
            ]
            soffice_cmd = None
            for path in possible_paths:
                if os.path.exists(path) or path == 'soffice.exe':
                    soffice_cmd = path
                    break
        else:
            # Linux/Mac에서는 일반적으로 PATH에 있음
            soffice_cmd = 'libreoffice'

        if not soffice_cmd:
            print("LibreOffice를 찾을 수 없습니다. PDF 변환을 건너뜁니다.")
            return False

        # LibreOffice headless 모드로 PDF 변환
        # --headless: GUI 없이 실행
        # --convert-to pdf: PDF로 변환
        # --outdir: 출력 디렉토리 지정
        cmd = [
            soffice_cmd,
            '--headless',
            '--convert-to',
            'pdf',
            '--outdir',
            output_dir,
            abs_excel_path
        ]

        print(f"PDF 변환 명령: {' '.join(cmd)}")

        # 프로세스 실행 (타임아웃 30초)
        result = subprocess.run(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            timeout=30,
            check=False
        )

        if result.returncode == 0:
            # LibreOffice는 파일명을 자동으로 .pdf로 바꿔서 저장
            # 원본 파일명에서 확장자만 .pdf로 변경
            base_name = os.path.splitext(os.path.basename(abs_excel_path))[0]
            generated_pdf = os.path.join(output_dir, f'{base_name}.pdf')

            # 원하는 PDF 이름으로 변경
            if os.path.exists(generated_pdf) and generated_pdf != pdf_path:
                if os.path.exists(pdf_path):
                    os.remove(pdf_path)
                os.rename(generated_pdf, pdf_path)

            # 임시 파일 정리
            if temp_excel_path and os.path.exists(temp_excel_path):
                try:
                    os.unlink(temp_excel_path)
                    print(f"임시 파일 삭제: {temp_excel_path}")
                except:
                    pass

            print(f"PDF 변환 성공: {pdf_path}")
            return True
        else:
            print(f"PDF 변환 실패 (returncode: {result.returncode})")
            print(f"stdout: {result.stdout.decode('utf-8', errors='ignore')}")
            print(f"stderr: {result.stderr.decode('utf-8', errors='ignore')}")

            # 실패 시에도 임시 파일 정리
            if temp_excel_path and os.path.exists(temp_excel_path):
                try:
                    os.unlink(temp_excel_path)
                except:
                    pass
            return False

    except subprocess.TimeoutExpired:
        print("PDF 변환 타임아웃 (30초 초과)")
        if temp_excel_path and os.path.exists(temp_excel_path):
            try:
                os.unlink(temp_excel_path)
            except:
                pass
        return False
    except FileNotFoundError:
        print("LibreOffice를 찾을 수 없습니다. PDF 변환을 건너뜁니다.")
        print("우분투: sudo apt-get install libreoffice")
        print("Windows: https://www.libreoffice.org/download/download/")
        if temp_excel_path and os.path.exists(temp_excel_path):
            try:
                os.unlink(temp_excel_path)
            except:
                pass
        return False
    except Exception as e:
        print(f"PDF 변환 오류: {str(e)}")
        import traceback
        traceback.print_exc()
        if temp_excel_path and os.path.exists(temp_excel_path):
            try:
                os.unlink(temp_excel_path)
            except:
                pass
        return False

@invoice_generator_bp.route('/generate-invoice', methods=['POST'])
def generate_invoice():
    """거래명세서 생성 API"""
    try:
        data = request.get_json()

        customer_name = data.get('customer_name')
        service_date = data.get('service_date')  # YYYY-MM-DD
        items = data.get('items', [])
        customer_info = data.get('customer_info', {})  # 고객사 정보

        if not customer_name or not service_date or not items:
            return jsonify({
                'success': False,
                'error': '필수 정보가 누락되었습니다.'
            }), 400

        # 날짜를 yymmdd 형식으로 변환
        date_obj = datetime.strptime(service_date, '%Y-%m-%d')
        sheet_name = date_obj.strftime('%y%m%d')

        # 템플릿 파일 전체를 복사하여 고객사별 파일 생성 (서식 100% 보존)
        workbook, invoice_file_path = copy_template_file(customer_name)

        # Template 시트들을 복사하여 날짜별 시트 생성 (고객사용, 공급자용)
        customer_sheet, supplier_sheet = prepare_sheet_from_template(workbook, sheet_name)

        # 합계금액 계산
        total_amount = calculate_total_amount(items)

        # 두 시트 모두에 데이터 입력 (동일한 셀 주소)
        for sheet in [customer_sheet, supplier_sheet]:
            # 헤더 정보 입력 (작성일자, 고객사명, 주소, 전화/팩스, 합계금액)
            write_header_info_to_sheet(sheet, service_date, customer_info, total_amount)

            # 데이터 입력
            write_invoice_items_to_sheet(sheet, items)

            # B3:AG43 영역에 두꺼운 테두리 적용
            apply_thick_border_to_range(sheet, 'B3', 'AG43')

        # 저장
        workbook.save(invoice_file_path)
        workbook.close()

        # PDF 변환 (순수 Python 방식 우선, LibreOffice는 백업)
        pdf_filename = f'거래명세표({customer_name})-{sheet_name}.pdf'
        pdf_path = os.path.join(os.path.dirname(invoice_file_path), pdf_filename)

        pdf_success = False
        pdf_url = None

        # 1순위: WeasyPrint (순수 Python, 외부 프로그램 불필요)
        if HAS_WEASYPRINT:
            pdf_success = convert_to_pdf_weasyprint(
                items, service_date, customer_info, total_amount, pdf_path
            )

        # 2순위: LibreOffice (WeasyPrint 실패 시)
        # 특정 시트만 PDF로 변환 (고객사용 시트만)
        if not pdf_success:
            customer_sheet_name_for_pdf = f"{sheet_name}-c"
            pdf_success = convert_excel_to_pdf_libreoffice(invoice_file_path, pdf_path, customer_sheet_name_for_pdf)

        if pdf_success:
            # 상대 경로 생성 (customer_name/filename)
            pdf_url = f'/api/invoice-pdf/{customer_name}/{pdf_filename}'

        # Excel 파일 다운로드 URL 추가
        excel_filename = f'거래명세표({customer_name}).xlsx'
        excel_url = f'/api/invoice-excel/{customer_name}/{excel_filename}'

        return jsonify({
            'success': True,
            'message': '거래명세표가 성공적으로 생성되었습니다.',
            'excel_path': invoice_file_path,
            'excel_url': excel_url,
            'pdf_path': pdf_path if pdf_success else None,
            'pdf_url': pdf_url,
            'sheet_name': sheet_name,
            'has_pdf': pdf_success
        })

    except Exception as e:
        import traceback
        error_traceback = traceback.format_exc()
        print(f"거래명세서 생성 오류:\n{error_traceback}")
        return jsonify({
            'success': False,
            'error': str(e),
            'details': error_traceback
        }), 500

@invoice_generator_bp.route('/invoice-excel/<customer_name>/<filename>', methods=['GET'])
def serve_invoice_excel(customer_name, filename):
    """생성된 Excel 파일 제공"""
    try:
        # instance/거래명세서/고객사명/filename 경로
        excel_path = os.path.join(INVOICE_BASE_DIR, customer_name, filename)
        if os.path.exists(excel_path):
            return send_file(
                excel_path,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
        else:
            return jsonify({
                'success': False,
                'error': 'Excel 파일을 찾을 수 없습니다.'
            }), 404
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@invoice_generator_bp.route('/invoice-pdf/<customer_name>/<filename>', methods=['GET'])
def serve_invoice_pdf(customer_name, filename):
    """생성된 PDF 파일 제공 (Excel이 설치된 경우에만)"""
    try:
        # instance/거래명세서/고객사명/filename 경로
        pdf_path = os.path.join(INVOICE_BASE_DIR, customer_name, filename)
        if os.path.exists(pdf_path):
            return send_file(pdf_path, mimetype='application/pdf')
        else:
            return jsonify({
                'success': False,
                'error': 'PDF 파일을 찾을 수 없습니다.'
            }), 404
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500
