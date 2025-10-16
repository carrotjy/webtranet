"""
템플릿 파일의 Template 시트를 Template-customer와 Template-supplier로 분리하는 스크립트
"""
from openpyxl import load_workbook
import os
import sys

# Windows 콘솔 인코딩 문제 해결
if sys.platform == 'win32':
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')

INSTANCE_DIR = os.path.join(os.path.dirname(__file__), 'instance')
TEMPLATE_PATH = os.path.join(INSTANCE_DIR, '거래명세표(양식).xlsx')

def split_template():
    """Template 시트를 Template-customer와 Template-supplier로 분리"""

    if not os.path.exists(TEMPLATE_PATH):
        print(f"❌ 템플릿 파일을 찾을 수 없습니다: {TEMPLATE_PATH}")
        return

    print(f"📂 템플릿 파일 열기: {TEMPLATE_PATH}")
    workbook = load_workbook(TEMPLATE_PATH)

    print(f"📋 현재 시트 목록: {workbook.sheetnames}")

    # Template 시트 찾기
    if 'Template' not in workbook.sheetnames:
        print("❌ 'Template' 시트를 찾을 수 없습니다.")
        print("   현재 시트:", workbook.sheetnames)
        workbook.close()
        return

    template_sheet = workbook['Template']

    # Template-customer 생성 (Template 시트 이름 변경)
    if 'Template-customer' not in workbook.sheetnames:
        print("✏️  'Template' 시트를 'Template-customer'로 이름 변경")
        template_sheet.title = 'Template-customer'
    else:
        print("✅ 'Template-customer' 시트가 이미 존재합니다.")

    # Template-supplier 생성 (Template-customer 복사)
    if 'Template-supplier' not in workbook.sheetnames:
        print("📋 'Template-customer' 시트를 복사하여 'Template-supplier' 생성")
        supplier_sheet = workbook.copy_worksheet(workbook['Template-customer'])
        supplier_sheet.title = 'Template-supplier'
    else:
        print("✅ 'Template-supplier' 시트가 이미 존재합니다.")

    # 저장
    print(f"💾 저장 중...")
    workbook.save(TEMPLATE_PATH)
    workbook.close()

    print("✅ 완료!")
    print(f"📋 최종 시트 목록:")

    # 저장된 파일 다시 열어서 확인
    wb_check = load_workbook(TEMPLATE_PATH)
    for sheet_name in wb_check.sheetnames:
        print(f"   - {sheet_name}")
    wb_check.close()

    print("\n⚠️  참고사항:")
    print("   - 두 템플릿 시트는 현재 동일한 내용입니다.")
    print("   - Excel에서 파일을 열어 각 템플릿의 수식/참조를 수정하세요.")
    print("   - Template-customer: 고객사용 양식")
    print("   - Template-supplier: 공급자용 양식")

if __name__ == '__main__':
    split_template()
