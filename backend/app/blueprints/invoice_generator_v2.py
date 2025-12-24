"""
신규 양식을 사용한 거래명세서 생성 (Name Define 방식)
"""
from flask import Blueprint, request, jsonify, send_file
import os
import shutil
import subprocess
from datetime import datetime
from openpyxl import load_workbook
from app.database.init_db import get_db_connection

# 상수
INSTANCE_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'instance')
INVOICE_BASE_DIR = os.path.join(INSTANCE_DIR, '거래명세서')
TEMPLATE_PATH = os.path.join(INSTANCE_DIR, '거래명세서_신규양식.xlsx')
NETWORK_BASE_DIR = '/mnt/windows/거래명세서'

def get_supplier_info():
    """공급자 정보 조회"""
    conn = get_db_connection()
    try:
        supplier = conn.execute('SELECT * FROM supplier_info ORDER BY id DESC LIMIT 1').fetchone()
        if supplier:
            return {
                'company_name': supplier['company_name'],
                'registration_number': supplier['registration_number'],
                'ceo_name': supplier['ceo_name'],
                'address': supplier['address'],
                'phone': supplier['phone'],
                'fax': supplier['fax']
            }
    finally:
        conn.close()

    return {
        'company_name': '',
        'registration_number': '',
        'ceo_name': '',
        'address': '',
        'phone': '',
        'fax': ''
    }

def get_libreoffice_path_from_settings():
    """시스템 설정에서 LibreOffice 경로 조회"""
    try:
        import sqlite3
        # webtranet.db에 연결 (system_settings 테이블이 있는 DB)
        db_path = os.path.join('app', 'database', 'webtranet.db')
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row

        setting = conn.execute(
            "SELECT value FROM system_settings WHERE key = 'libreoffice_path'"
        ).fetchone()
        conn.close()

        if setting and setting['value']:
            print(f"✅ DB에서 LibreOffice 경로 조회 성공: {setting['value']}")
            return setting['value']
        else:
            print("❌ DB에 LibreOffice 경로가 저장되지 않음")
        return None
    except Exception as e:
        print(f"❌ LibreOffice 경로 설정 조회 실패: {str(e)}")
        return None

def convert_excel_to_pdf(excel_path, pdf_path):
    """
    LibreOffice를 사용하여 Excel을 PDF로 변환

    Args:
        excel_path: Excel 파일 경로
        pdf_path: 생성할 PDF 파일 경로

    Returns:
        bool: 성공 여부
    """
    try:
        import platform
        output_dir = os.path.dirname(pdf_path)

        # 1순위: 시스템 설정에서 LibreOffice 경로 확인
        soffice_cmd = None
        custom_path = get_libreoffice_path_from_settings()
        if custom_path:
            print(f"시스템 설정에서 LibreOffice 경로 발견: {custom_path}")
            if os.path.exists(custom_path):
                soffice_cmd = custom_path
                print(f"사용자 지정 LibreOffice 경로 사용: {soffice_cmd}")
            else:
                print(f"경고: 설정된 경로가 존재하지 않음: {custom_path}")

        # 2순위: 기본 경로 확인
        if not soffice_cmd:
            system = platform.system()
            if system == 'Windows':
                possible_paths = [
                    r'C:\Program Files\LibreOffice\program\soffice.exe',
                    r'C:\Program Files (x86)\LibreOffice\program\soffice.exe',
                ]
                for path in possible_paths:
                    if os.path.exists(path):
                        soffice_cmd = path
                        print(f"기본 경로에서 LibreOffice 찾음: {path}")
                        break
                if not soffice_cmd:
                    soffice_cmd = 'libreoffice'  # PATH에서 시도
            else:
                soffice_cmd = 'libreoffice'  # Linux/Mac

        if not soffice_cmd:
            print("❌ LibreOffice를 찾을 수 없습니다")
            return False

        print(f"최종 LibreOffice 명령어: {soffice_cmd}")

        # LibreOffice 명령어 실행
        cmd = [
            soffice_cmd,
            '--headless',
            '--convert-to', 'pdf',
            '--outdir', output_dir,
            excel_path
        ]

        print(f"PDF 변환 시작: {excel_path} -> {pdf_path}")
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)

        if result.returncode == 0:
            # LibreOffice는 원본 파일명을 기반으로 PDF를 생성하므로
            # 원하는 파일명과 다른 경우 rename 필요
            excel_basename = os.path.splitext(os.path.basename(excel_path))[0]
            generated_pdf = os.path.join(output_dir, f"{excel_basename}.pdf")

            # 생성된 PDF 파일명이 원하는 파일명과 다르면 rename
            if generated_pdf != pdf_path and os.path.exists(generated_pdf):
                import shutil
                shutil.move(generated_pdf, pdf_path)
                print(f"✅ PDF 파일명 변경: {generated_pdf} -> {pdf_path}")

            if os.path.exists(pdf_path):
                print(f"✅ PDF 변환 성공: {pdf_path}")
                return True
            else:
                print(f"❌ PDF 파일을 찾을 수 없음: {pdf_path}")
                return False
        else:
            print(f"❌ PDF 변환 실패: {result.stderr}")
            return False

    except subprocess.TimeoutExpired:
        print("❌ PDF 변환 시간 초과")
        return False
    except FileNotFoundError:
        print("❌ LibreOffice가 설치되지 않음")
        return False
    except Exception as e:
        print(f"❌ PDF 변환 오류: {str(e)}")
        return False

def write_value_by_name(workbook, name, value):
    """Name Define을 사용하여 값 쓰기 (merged cell 처리 포함)"""
    try:
        if name not in workbook.defined_names:
            print(f"Warning: Name '{name}' not found in workbook")
            return False

        # Name Define에서 셀 위치 가져오기
        for title, coord in workbook.defined_names[name].destinations:
            sheet = workbook[title]
            cell = sheet[coord]

            # MergedCell인 경우 merged range의 top-left cell에 쓰기
            if hasattr(cell, '__class__') and 'MergedCell' in str(type(cell)):
                # Find the merged range containing this cell
                for merged_range in sheet.merged_cells.ranges:
                    if coord in merged_range:
                        # Write to the top-left cell of the merged range
                        top_left = merged_range.start_cell.coordinate
                        sheet[top_left] = value
                        return True
                # If not in any merged range, try writing directly (should fail)
                print(f"Warning: Cell {coord} is MergedCell but not in any merged range")
                return False
            else:
                # Normal cell - write directly
                sheet[coord] = value
                return True
    except Exception as e:
        print(f"Error writing to name '{name}': {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def generate_invoice_excel_v2(invoice_id):
    """
    DB의 invoice 데이터를 기반으로 Excel 파일 생성 (Name Define 방식)

    Args:
        invoice_id: 거래명세서 ID

    Returns:
        dict: 생성 결과 {'success': bool, 'file_path': str, 'message': str}
    """
    conn = get_db_connection()
    try:
        # 1. Invoice 데이터 조회
        invoice = conn.execute('''
            SELECT * FROM invoices WHERE id = ?
        ''', (invoice_id,)).fetchone()

        if not invoice:
            return {'success': False, 'message': '거래명세서를 찾을 수 없습니다.'}

        # 디버깅: invoice_number 확인
        print(f"🔍 Invoice ID {invoice_id} 조회: invoice_number = '{invoice['invoice_number']}'")
        print(f"   customer_name = '{invoice['customer_name']}'")
        print(f"   issue_date = '{invoice['issue_date']}'")

        # 2. Invoice 항목 조회 (row_order 순서로)
        items = conn.execute('''
            SELECT * FROM invoice_items
            WHERE invoice_id = ?
            ORDER BY row_order, id
        ''', (invoice_id,)).fetchall()

        # 3. 고객사 정보 조회 (customers 테이블에서)
        customer = None
        if invoice['customer_id']:
            customer = conn.execute('''
                SELECT * FROM customers WHERE id = ?
            ''', (invoice['customer_id'],)).fetchone()

        # 4. 공급자 정보 조회
        supplier = get_supplier_info()

        # 5. 템플릿 복사
        customer_folder = os.path.join(INVOICE_BASE_DIR, invoice['customer_name'])
        os.makedirs(customer_folder, exist_ok=True)

        output_filename = f"거래명세서({invoice['customer_name']}).xlsx"
        output_path = os.path.join(customer_folder, output_filename)

        # 템플릿 복사
        shutil.copy2(TEMPLATE_PATH, output_path)

        # 6. Excel 파일 열기
        wb = load_workbook(output_path)

        # 7. 공급자 정보 입력
        write_value_by_name(wb, 'provider_name', supplier['company_name'])
        write_value_by_name(wb, 'provider_president', supplier['ceo_name'])
        write_value_by_name(wb, 'provider_address', supplier['address'])
        write_value_by_name(wb, 'provider_number', supplier['registration_number'])
        write_value_by_name(wb, 'provider_tel', supplier['phone'])
        write_value_by_name(wb, 'provider_fax', supplier['fax'])

        # 8. 고객사 정보 입력
        write_value_by_name(wb, 'customer_name', invoice['customer_name'])
        write_value_by_name(wb, 'customer_address', invoice['customer_address'] or (customer['address'] if customer else ''))
        write_value_by_name(wb, 'customer_tel', invoice['customer_tel'] or (customer['phone'] if customer else ''))
        write_value_by_name(wb, 'customer_fax', invoice['customer_fax'] or (customer['fax'] if customer else ''))

        # 9. 기타 정보 입력
        write_value_by_name(wb, 'invoice_number', invoice['invoice_number'])
        write_value_by_name(wb, 'issue_date', invoice['issue_date'])

        # 10. 금액 정보 입력
        write_value_by_name(wb, 'amount_price', invoice['total_amount'])
        write_value_by_name(wb, 'tax_price', invoice['vat_amount'])
        write_value_by_name(wb, 'total_amount', invoice['grand_total'])

        # 11. 항목 데이터 입력 (16행부터 시작 - 15행은 헤더)
        sheet = wb.active
        current_row = 16  # 시작 행 (15행은 헤더 행)

        # 34행 이후 추가 행이 필요한 경우를 위한 템플릿 행 준비
        template_row = 16  # 16행을 템플릿으로 사용

        # 병합된 셀 안전 처리 함수
        def safe_write_to_cell(cell_ref, value):
            """병합된 셀을 안전하게 처리하며 값 입력"""
            try:
                # 병합된 셀 범위 확인
                for merged_range in sheet.merged_cells.ranges:
                    if cell_ref in merged_range:
                        # 병합 범위의 왼쪽 상단 셀 좌표 가져오기
                        min_col, min_row, max_col, max_row = merged_range.bounds
                        top_left_cell = sheet.cell(row=min_row, column=min_col)
                        top_left_cell.value = value
                        return top_left_cell
                # 병합되지 않은 셀이면 직접 입력
                sheet[cell_ref].value = value
                return sheet[cell_ref]
            except Exception as e:
                print(f"셀 {cell_ref}에 값 입력 실패: {str(e)}")
                return None

        for item in items:
            # 34행을 초과하는 경우 새 행 삽입 및 스타일 복사
            if current_row > 34:
                from openpyxl.worksheet.cell_range import CellRange
                from openpyxl.utils import get_column_letter

                sheet.insert_rows(current_row)

                # 템플릿 행의 스타일 복사
                for col in range(1, sheet.max_column + 1):
                    source_cell = sheet.cell(row=template_row, column=col)
                    target_cell = sheet.cell(row=current_row, column=col)
                    if source_cell.has_style:
                        target_cell.font = source_cell.font.copy()
                        target_cell.border = source_cell.border.copy()
                        target_cell.fill = source_cell.fill.copy()
                        target_cell.number_format = source_cell.number_format
                        target_cell.alignment = source_cell.alignment.copy()

                # 병합된 셀 복사 (C:H, I:N, O:P, Q:V, W:AB, AC:AG)
                merge_ranges = [
                    (3, 8),   # C:H (품목)
                    (9, 14),  # I:N (규격)
                    (15, 16), # O:P (수량)
                    (17, 22), # Q:V (단가)
                    (23, 28), # W:AB (공급가액)
                    (29, 33)  # AC:AG (세액)
                ]
                for start_col, end_col in merge_ranges:
                    sheet.merge_cells(
                        start_row=current_row,
                        start_column=start_col,
                        end_row=current_row,
                        end_column=end_col
                    )

            # 헤더 행은 월, 일, 품목만
            if item['is_header']:
                if item['month']:
                    safe_write_to_cell(f'A{current_row}', item['month'])
                if item['day']:
                    safe_write_to_cell(f'B{current_row}', item['day'])
                if item['item_name']:
                    # 품목은 C:H merged range의 시작 셀인 C에 쓰기
                    header_cell = safe_write_to_cell(f'C{current_row}', item['item_name'])
                    # 헤더 행도 자동 줄바꿈 설정
                    if header_cell:
                        from openpyxl.styles import Alignment
                        old_alignment = header_cell.alignment
                        header_cell.alignment = Alignment(
                            wrap_text=True,
                            horizontal=old_alignment.horizontal if old_alignment else 'left',
                            vertical=old_alignment.vertical if old_alignment else 'center'
                        )
                current_row += 1
                continue

            # 일반 행
            # 월/일은 헤더 행에만 표시 (일반 행에는 입력하지 않음)
            # if item['month']:
            #     safe_write_to_cell(f'A{current_row}', item['month'])
            # if item['day']:
            #     safe_write_to_cell(f'B{current_row}', item['day'])

            # 네고 항목 판별 (음수 금액 또는 "네고" 포함)
            is_nego = item['total_price'] < 0 if item['total_price'] else False
            if not is_nego and item['item_name']:
                is_nego = '네고' in str(item['item_name']) or 'NEGO' in str(item['item_name']).upper()

            if item['item_name']:
                # 품목은 C에 쓰기
                item_name = str(item['item_name']).replace('네고', 'NEGO')
                item_cell = safe_write_to_cell(f'C{current_row}', item_name)

                # 자동 줄바꿈 설정 (긴 품목명 처리)
                if item_cell:
                    from openpyxl.styles import Alignment, Font
                    old_alignment = item_cell.alignment
                    item_cell.alignment = Alignment(
                        wrap_text=True,
                        horizontal=old_alignment.horizontal if old_alignment else 'left',
                        vertical=old_alignment.vertical if old_alignment else 'center'
                    )

                    # 네고 항목은 빨간색으로 표시 (기존 폰트 속성 유지)
                    if is_nego:
                        old_font = item_cell.font
                        item_cell.font = Font(
                            name=old_font.name,
                            size=old_font.size,
                            bold=old_font.bold,
                            italic=old_font.italic,
                            vertAlign=old_font.vertAlign,
                            underline=old_font.underline,
                            strike=old_font.strike,
                            color="FF0000"
                        )

            if item['description']:
                # 규격은 J에 쓰기
                desc_cell = safe_write_to_cell(f'J{current_row}', item['description'])
                if is_nego and desc_cell:
                    from openpyxl.styles import Font
                    old_font = desc_cell.font
                    desc_cell.font = Font(
                        name=old_font.name,
                        size=old_font.size,
                        bold=old_font.bold,
                        italic=old_font.italic,
                        vertAlign=old_font.vertAlign,
                        underline=old_font.underline,
                        strike=old_font.strike,
                        color="FF0000"
                    )

            if item['quantity']:
                # 수량은 Q에 쓰기
                qty_cell = safe_write_to_cell(f'Q{current_row}', item['quantity'])

                if qty_cell:
                    # 품목 타입에 따라 셀 서식 적용
                    if item['item_type'] == 'work' or item['item_type'] == 'travel':
                        # 작업시간/이동시간: 소수점 1자리 + "H" 표시
                        qty_cell.number_format = '0.0"H"'
                    else:
                        # 부품: "EA" 표시
                        qty_cell.number_format = '0"EA"'

                    # 네고 항목은 빨간색으로 표시 (기존 폰트 속성 유지)
                    if is_nego:
                        from openpyxl.styles import Font
                        old_font = qty_cell.font
                        qty_cell.font = Font(
                            name=old_font.name,
                            size=old_font.size,
                            bold=old_font.bold,
                            italic=old_font.italic,
                            vertAlign=old_font.vertAlign,
                            underline=old_font.underline,
                            strike=old_font.strike,
                            color="FF0000"
                        )

            if item['unit_price']:
                # 단가는 S에 쓰기
                price_cell = safe_write_to_cell(f'S{current_row}', item['unit_price'])
                if is_nego and price_cell:
                    from openpyxl.styles import Font
                    old_font = price_cell.font
                    price_cell.font = Font(
                        name=old_font.name,
                        size=old_font.size,
                        bold=old_font.bold,
                        italic=old_font.italic,
                        vertAlign=old_font.vertAlign,
                        underline=old_font.underline,
                        strike=old_font.strike,
                        color="FF0000"
                    )

            if item['total_price']:
                # 공급가액은 X에 쓰기
                total_cell = safe_write_to_cell(f'X{current_row}', item['total_price'])
                if is_nego and total_cell:
                    from openpyxl.styles import Font
                    old_font = total_cell.font
                    total_cell.font = Font(
                        name=old_font.name,
                        size=old_font.size,
                        bold=old_font.bold,
                        italic=old_font.italic,
                        vertAlign=old_font.vertAlign,
                        underline=old_font.underline,
                        strike=old_font.strike,
                        color="FF0000"
                    )

            # 세액 계산 (10%)
            vat = round(item['total_price'] * 0.1) if item['total_price'] else 0
            if vat:
                # 세액은 AC에 쓰기
                vat_cell = safe_write_to_cell(f'AC{current_row}', vat)
                if is_nego and vat_cell:
                    from openpyxl.styles import Font
                    old_font = vat_cell.font
                    vat_cell.font = Font(
                        name=old_font.name,
                        size=old_font.size,
                        bold=old_font.bold,
                        italic=old_font.italic,
                        vertAlign=old_font.vertAlign,
                        underline=old_font.underline,
                        strike=old_font.strike,
                        color="FF0000"
                    )

            current_row += 1

        # 12. 행 높이 자동 조정 (줄바꿈된 셀 대응)
        # 16행부터 데이터가 입력된 마지막 행까지 처리
        for row_num in range(16, current_row):
            # 각 행의 최소 높이 유지 (기본 템플릿 높이)
            # wrap_text가 설정된 셀이 있는 경우 자동으로 높이 조정됨
            # openpyxl은 자동 높이 계산을 지원하지 않으므로,
            # 최소 높이만 설정하고 LibreOffice PDF 변환 시 자동 조정됨
            if row_num not in sheet.row_dimensions:
                sheet.row_dimensions[row_num].height = None  # 자동 높이
            elif sheet.row_dimensions[row_num].height:
                # 기존 높이가 있으면 최소 높이로 설정 (필요시 늘어남)
                pass

        # 13. 저장
        wb.save(output_path)
        wb.close()

        # 14. PDF 생성 (LibreOffice 사용) - 월별 폴더에 저장
        # 발행일자에서 년월 추출 (YYYY-MM-DD 형식)
        issue_date_str = invoice['issue_date']
        try:
            issue_date = datetime.strptime(issue_date_str, '%Y-%m-%d')
            monthly_folder_name = f"{issue_date.year}년{issue_date.month:02d}월"
        except:
            # 날짜 파싱 실패 시 현재 날짜 사용
            now = datetime.now()
            monthly_folder_name = f"{now.year}년{now.month:02d}월"

        # 월별 폴더 생성
        monthly_folder = os.path.join(INVOICE_BASE_DIR, monthly_folder_name)
        os.makedirs(monthly_folder, exist_ok=True)

        # PDF 파일명 생성 (invoice_number가 없으면 고객명만 사용)
        invoice_number = invoice['invoice_number']
        if invoice_number:
            pdf_filename = f"거래명세서({invoice['customer_name']})-{invoice_number}.pdf"
            print(f"✅ PDF 파일명: {pdf_filename}")
        else:
            print(f"⚠️  경고: Invoice ID {invoice_id}의 invoice_number가 비어있습니다!")
            pdf_filename = f"거래명세서({invoice['customer_name']}).pdf"
            print(f"   기본 PDF 파일명 사용: {pdf_filename}")

        pdf_path = os.path.join(monthly_folder, pdf_filename)
        print(f"📁 PDF 저장 경로: {pdf_path}")
        pdf_success = convert_excel_to_pdf(output_path, pdf_path)

        return {
            'success': True,
            'file_path': output_path,
            'filename': output_filename,
            'pdf_path': pdf_path if pdf_success else None,
            'pdf_filename': pdf_filename if pdf_success else None,
            'message': '거래명세서가 성공적으로 생성되었습니다.' + (' (PDF 포함)' if pdf_success else ' (Excel만)')
        }

    except Exception as e:
        import traceback
        traceback.print_exc()
        return {
            'success': False,
            'message': f'거래명세서 생성 실패: {str(e)}'
        }
    finally:
        conn.close()
