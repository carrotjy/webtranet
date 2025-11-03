from flask import Blueprint, request, jsonify, send_file
import sqlite3
import os
from datetime import datetime, date
from flask_jwt_extended import jwt_required, get_jwt_identity
from app.models.user import User
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

spare_parts_bp = Blueprint('spare_parts', __name__)

def get_db_connection():
    """데이터베이스 연결"""
    db_path = os.path.join('app', 'database', 'user.db')
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn

@spare_parts_bp.route('/spare-parts', methods=['GET'])
def get_spare_parts():
    """스페어파트 목록 조회"""
    try:
        conn = get_db_connection()
        
        # 검색 파라미터 가져오기
        search_term = request.args.get('search', '').strip()
        
        # 기본 쿼리
        if search_term:
            # 파트번호 또는 파트명으로 검색 (부분 일치)
            query = '''
                SELECT * FROM spare_parts 
                WHERE part_number LIKE ? OR part_name LIKE ? 
                ORDER BY part_number
            '''
            search_pattern = f'%{search_term}%'
            parts = conn.execute(query, (search_pattern, search_pattern)).fetchall()
        else:
            # 검색어가 없으면 전체 조회
            parts = conn.execute('SELECT * FROM spare_parts ORDER BY part_number').fetchall()
            
        # 결과 변환 - 프론트엔드 형식에 맞춤 (최신 청구가 포함)
        parts_list = []
        for part in parts:
            # 각 부품의 최신 청구가 조회
            latest_billing_price = conn.execute('''
                SELECT billing_price 
                FROM price_history 
                WHERE spare_part_id = ? 
                ORDER BY effective_date DESC, created_at DESC 
                LIMIT 1
            ''', (part['id'],)).fetchone()
            
            billing_price = latest_billing_price['billing_price'] if latest_billing_price and latest_billing_price['billing_price'] else 0
            
            parts_list.append({
                'id': part['id'],  # 실제 id 필드 사용
                'part_number': part['part_number'],
                'part_name': part['part_name'],  # part_name 필드명 유지
                'erp_name': part['erp_name'] if part['erp_name'] else '',  # erp_name 추가
                'stock_quantity': part['stock_quantity'],  # stock_quantity 필드명 유지
                'price': part['price'] if part['price'] else 0,  # price 필드명 유지
                'billing_price': billing_price,  # 최신 청구가 추가
                'created_at': part['created_at'] if part['created_at'] else datetime.now().isoformat(),
                'updated_at': part['updated_at'] if part['updated_at'] else datetime.now().isoformat()
            })
        
        conn.close()
        
        return jsonify(parts_list)
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@spare_parts_bp.route('/spare-parts', methods=['POST'])
def create_spare_part():
    """새 스페어파트 생성"""
    try:
        data = request.get_json()
        
        part_number = data.get('part_number')
        part_name = data.get('part_name')
        erp_name = data.get('erp_name', '')
        description = data.get('description', '')
        category = data.get('category', '')
        current_stock = data.get('stock_quantity', 0)  # 프론트엔드 필드명에 맞춤
        min_stock = data.get('min_stock', 0)
        price_eur = data.get('price', 0)  # 프론트엔드 필드명에 맞춤
        
        print(f"Creating new part with data: {data}")  # 디버깅용
        
        if not part_number or not part_name:
            return jsonify({
                'success': False,
                'error': '파트번호와 파트명은 필수입니다.'
            }), 400
        
        conn = get_db_connection()
        
        # 중복 체크
        existing = conn.execute(
            'SELECT part_number FROM spare_parts WHERE part_number = ?',
            (part_number,)
        ).fetchone()
        
        if existing:
            conn.close()
            return jsonify({
                'success': False,
                'error': '이미 존재하는 파트번호입니다.'
            }), 400
        
        # 삽입 - erp_name 필드 추가
        conn.execute('''
            INSERT INTO spare_parts 
            (part_number, part_name, erp_name, description, price, stock_quantity, minimum_stock, supplier)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (part_number, part_name, erp_name, description, price_eur, current_stock, min_stock, category))
        
        conn.commit()
        
        # 삽입 후 확인
        inserted_part = conn.execute(
            'SELECT * FROM spare_parts WHERE part_number = ?', 
            (part_number,)
        ).fetchone()
        
        print(f"Successfully created part {part_number} with erp_name: {erp_name}")  # 디버깅용
        print(f"Inserted part data: {dict(inserted_part) if inserted_part else 'None'}")  # 디버깅용
        
        conn.close()
        
        return jsonify({
            'success': True,
            'message': '스페어파트가 생성되었습니다.'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@spare_parts_bp.route('/spare-parts/<part_number>', methods=['PUT'])
def update_spare_part(part_number):
    """스페어파트 수정"""
    try:
        data = request.get_json()
        
        part_name = data.get('part_name')
        erp_name = data.get('erp_name', '')
        description = data.get('description', '')
        category = data.get('category', '')
        current_stock = data.get('stock_quantity', 0)  # 프론트엔드 필드명에 맞춤
        min_stock = data.get('min_stock', 0)
        price_eur = data.get('price', 0)  # 프론트엔드 필드명에 맞춤
        
        print(f"Updating part {part_number} with data: {data}")  # 디버깅용
        
        if not part_name:
            return jsonify({
                'success': False,
                'error': '파트명은 필수입니다.'
            }), 400
        
        conn = get_db_connection()
        
        # 업데이트 - erp_name 필드 추가
        conn.execute('''
            UPDATE spare_parts 
            SET part_name = ?, erp_name = ?, description = ?, price = ?, 
                stock_quantity = ?, minimum_stock = ?, supplier = ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE part_number = ?
        ''', (part_name, erp_name, description, price_eur, current_stock, min_stock, category, part_number))
        
        if conn.total_changes == 0:
            conn.close()
            return jsonify({
                'success': False,
                'error': '해당 부품을 찾을 수 없습니다.'
            }), 404
        
        print(f"Successfully updated part {part_number} with erp_name: {erp_name}")  # 디버깅용
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': '스페어파트가 수정되었습니다.'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@spare_parts_bp.route('/spare-parts/<int:part_id>', methods=['DELETE'])
def delete_spare_part(part_id):
    """스페어파트 삭제"""
    try:
        conn = get_db_connection()
        
        # 삭제
        conn.execute('DELETE FROM spare_parts WHERE id = ?', (part_id,))
        
        if conn.total_changes == 0:
            conn.close()
            return jsonify({
                'success': False,
                'error': '해당 부품을 찾을 수 없습니다.'
            }), 404
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': '스페어파트가 삭제되었습니다.'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@spare_parts_bp.route('/spare-parts/inbound', methods=['POST'])
def process_inbound():
    """입고 처리"""
    try:
        data = request.get_json()
        part_number = data.get('part_number')
        part_name = data.get('part_name')
        quantity = data.get('quantity', 0)
        currency = data.get('currency', 'EUR')
        price = data.get('price', 0)
        transaction_date = data.get('transaction_date')
        
        if not all([part_number, part_name, quantity > 0]):
            return jsonify({
                'success': False,
                'error': '필수 필드가 누락되었습니다.'
            }), 400
        
        conn = get_db_connection()
        
        # 기존 파트 확인
        existing_part = conn.execute(
            'SELECT * FROM spare_parts WHERE part_number = ?', 
            (part_number,)
        ).fetchone()
        
        if existing_part:
            # 기존 파트인 경우 재고 업데이트
            new_stock = existing_part['stock_quantity'] + quantity
            
            # 가격 정보 업데이트 (EUR 기준으로 저장)
            if currency == 'EUR':
                updated_price = price
            elif currency == 'USD':
                updated_price = price * 0.85  # USD to EUR 환율 (임시)
            else:  # KRW
                updated_price = price / 1320  # KRW to EUR 환율 (임시)
            
            conn.execute('''
                UPDATE spare_parts 
                SET stock_quantity = ?, price = ?
                WHERE part_number = ?
            ''', (new_stock, updated_price, part_number))
            
            # stock_history에 입고 기록 추가
            conn.execute('''
                INSERT INTO stock_history 
                (part_number, transaction_type, quantity, previous_stock, new_stock, 
                 transaction_date, created_at, created_by)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (part_number, 'IN', quantity, existing_part['stock_quantity'], 
                  new_stock, transaction_date or datetime.now().date(), 
                  datetime.now(), 'system'))
            
        else:
            # 신규 파트 등록
            if currency == 'EUR':
                eur_price = price
            elif currency == 'USD':
                eur_price = price * 0.85
            else:  # KRW
                eur_price = price / 1320
            
            conn.execute('''
                INSERT INTO spare_parts 
                (part_number, part_name, description, supplier, stock_quantity, minimum_stock, price)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (part_number, part_name, '', 'Unknown', quantity, 0, eur_price))
            
            # stock_history에 입고 기록 추가 (신규 파트)
            conn.execute('''
                INSERT INTO stock_history 
                (part_number, transaction_type, quantity, previous_stock, new_stock, 
                 transaction_date, created_at, created_by)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (part_number, 'IN', quantity, 0, quantity, 
                  transaction_date or datetime.now().date(), 
                  datetime.now(), 'system'))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': '입고 처리가 완료되었습니다.'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@spare_parts_bp.route('/spare-parts/outbound', methods=['POST'])
def process_outbound():
    """출고 처리"""
    try:
        data = request.get_json()
        part_number = data.get('part_number')
        quantity = data.get('quantity', 0)
        
        if not all([part_number, quantity > 0]):
            return jsonify({
                'success': False,
                'error': '필수 필드가 누락되었습니다.'
            }), 400
        
        conn = get_db_connection()
        
        # 기존 파트 확인
        existing_part = conn.execute(
            'SELECT * FROM spare_parts WHERE part_number = ?', 
            (part_number,)
        ).fetchone()
        
        if not existing_part:
            conn.close()
            return jsonify({
                'success': False,
                'error': '등록되지 않은 파트입니다.'
            }), 404
        
        if existing_part['stock_quantity'] < quantity:
            conn.close()
            return jsonify({
                'success': False,
                'error': '재고가 부족합니다.'
            }), 400
        
        # 재고 차감
        new_stock = existing_part['stock_quantity'] - quantity
        conn.execute('''
            UPDATE spare_parts 
            SET stock_quantity = ?
            WHERE part_number = ?
        ''', (new_stock, part_number))
        
        # stock_history에 출고 기록 추가
        conn.execute('''
            INSERT INTO stock_history 
            (part_number, transaction_type, quantity, previous_stock, new_stock, 
             transaction_date, created_at, created_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (part_number, 'OUT', quantity, existing_part['stock_quantity'], 
              new_stock, datetime.now().date(), datetime.now(), 'system'))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': '출고 처리가 완료되었습니다.'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@spare_parts_bp.route('/spare-parts/search/<part_number>', methods=['GET'])
def search_part_by_number(part_number):
    """파트번호로 검색"""
    try:
        conn = get_db_connection()
        
        part = conn.execute(
            'SELECT * FROM spare_parts WHERE part_number = ?', 
            (part_number,)
        ).fetchone()
        
        conn.close()
        
        if part:
            return jsonify({
                'success': True,
                'data': {
                    'part_number': part['part_number'],
                    'part_name': part['part_name'],
                    'description': part['description'],
                    'category': part['supplier'],
                    'current_stock': part['stock_quantity'],
                    'min_stock': part['minimum_stock'],
                    'current_price_eur': part['price'],
                    'current_price_krw': part['price'] * 1320 if part['price'] else None
                }
            })
        else:
            return jsonify({
                'success': False,
                'error': '파트를 찾을 수 없습니다.'
            }), 404
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@spare_parts_bp.route('/transactions', methods=['GET'])
def get_all_transactions():
    """모든 거래 내역 조회"""
    try:
        conn = get_db_connection()
        
        # 페이지네이션 파라미터
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 20))
        offset = (page - 1) * per_page
        
        # 필터 파라미터
        part_number_filter = request.args.get('part_number', '')
        transaction_type_filter = request.args.get('transaction_type', '')
        days_filter = request.args.get('days', '')
        
        # 기본 쿼리
        where_conditions = []
        params = []
        
        # 파트번호 필터
        if part_number_filter:
            where_conditions.append("sh.part_number LIKE ?")
            params.append(f'%{part_number_filter}%')
        
        # 거래유형 필터
        if transaction_type_filter:
            if transaction_type_filter == 'inbound':
                where_conditions.append("sh.transaction_type = ?")
                params.append('IN')
            elif transaction_type_filter == 'outbound':
                where_conditions.append("sh.transaction_type = ?")
                params.append('OUT')
        
        # 기간 필터
        if days_filter and days_filter != 'all':
            days = int(days_filter)
            where_conditions.append("sh.transaction_date >= DATE('now', '-{} days')".format(days))
        
        where_clause = "WHERE " + " AND ".join(where_conditions) if where_conditions else ""
        
        # 총 개수 조회
        count_query = f"""
            SELECT COUNT(*) as total
            FROM stock_history sh
            JOIN spare_parts sp ON sh.part_number = sp.part_number
            {where_clause}
        """
        
        total = conn.execute(count_query, params).fetchone()['total']
        
        # 데이터 조회
        data_query = f"""
            SELECT 
                sh.id,
                sh.part_number,
                sh.transaction_type,
                sh.quantity as quantity_change,
                sh.new_stock as current_quantity,
                sh.transaction_date,
                sh.created_at,
                sh.customer_name,
                sh.reference_number,
                sh.created_by,
                sp.part_name,
                sp.price as unit_price
            FROM stock_history sh
            JOIN spare_parts sp ON sh.part_number = sp.part_number
            {where_clause}
            ORDER BY sh.created_at DESC
            LIMIT ? OFFSET ?
        """
        
        params.extend([per_page, offset])
        transactions_raw = conn.execute(data_query, params).fetchall()
        conn.close()
        
        # 결과 포맷팅
        transactions = []
        for t in transactions_raw:
            total_amount = 0
            if t['transaction_type'] == 'IN' and t['unit_price']:
                total_amount = abs(t['quantity_change']) * t['unit_price']
            
            transactions.append({
                'id': t['id'],
                'date': t['created_at'] if t['created_at'] else t['transaction_date'],
                'transaction_date': t['transaction_date'],
                'transaction_type': 'inbound' if t['transaction_type'] == 'IN' else 'outbound',
                'part_number': t['part_number'],
                'part_name': t['part_name'],
                'quantity_change': t['quantity_change'],
                'quantity': abs(t['quantity_change']),
                'unit_price': t['unit_price'],
                'total_amount': total_amount,
                'current_quantity': t['current_quantity'],
                'new_stock': t['current_quantity'],
                'customer_name': t['customer_name'],
                'reference_number': t['reference_number'],
                'created_by': t['created_by']
            })
        
        pages = (total + per_page - 1) // per_page
        
        return jsonify({
            'success': True,
            'data': transactions,
            'pagination': {
                'page': page,
                'pages': pages,
                'per_page': per_page,
                'total': total
            }
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

@spare_parts_bp.route('/spare-parts/history', methods=['GET'])
def get_spare_parts_history():
    """스페어파트 입출고 내역 조회"""
    try:
        conn = get_db_connection()
        
        # 입출고 내역 조회 (최신순)
        history_records = conn.execute('''
            SELECT 
                sh.id,
                sh.part_number,
                sp.part_name,
                sh.transaction_type,
                sh.quantity,
                sh.previous_stock,
                sh.new_stock,
                sh.transaction_date,
                sh.created_at,
                sh.created_by,
                sh.reference_number,
                sh.customer_name
            FROM stock_history sh
            LEFT JOIN spare_parts sp ON sh.part_number = sp.part_number
            ORDER BY sh.created_at DESC
            LIMIT 100
        ''').fetchall()
        
        conn.close()
        
        # 결과 변환
        history_list = []
        for record in history_records:
            history_list.append({
                'id': record['id'],
                'part_number': record['part_number'],
                'part_name': record['part_name'] or 'Unknown',
                'transaction_type': record['transaction_type'],
                'quantity': record['quantity'],
                'previous_stock': record['previous_stock'],
                'new_stock': record['new_stock'],
                'transaction_date': record['transaction_date'] or record['created_at'][:10],
                'created_at': record['created_at'],
                'created_by': record['created_by'] or 'system',
                'reference_number': record['reference_number'] or '',
                'customer_name': record['customer_name'] or ''
            })
        
        return jsonify(history_list)
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@spare_parts_bp.route('/spare-parts/<int:part_id>/stock', methods=['POST'])
def update_stock(part_id):
    """부품 재고 입출고 처리"""
    try:
        data = request.get_json()
        
        transaction_type = data.get('type')  # 'in' 또는 'out'
        quantity = data.get('quantity', 0)
        
        if not transaction_type or quantity <= 0:
            return jsonify({
                'success': False,
                'error': '거래 유형과 수량이 필요합니다.'
            }), 400
        
        conn = get_db_connection()
        
        # 부품 정보 조회
        part = conn.execute('SELECT * FROM spare_parts WHERE id = ?', (part_id,)).fetchone()
        if not part:
            conn.close()
            return jsonify({
                'success': False,
                'error': '해당 부품을 찾을 수 없습니다.'
            }), 404
        
        current_stock = part['stock_quantity']
        
        if transaction_type == 'in':
            # 입고 처리
            new_stock = current_stock + quantity
        elif transaction_type == 'out':
            # 출고 처리
            if current_stock < quantity:
                conn.close()
                return jsonify({
                    'success': False,
                    'error': '재고가 부족합니다.'
                }), 400
            new_stock = current_stock - quantity
        else:
            conn.close()
            return jsonify({
                'success': False,
                'error': '잘못된 거래 유형입니다.'
            }), 400
        
        # 재고 업데이트
        conn.execute('''
            UPDATE spare_parts 
            SET stock_quantity = ?, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (new_stock, part_id))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'{"입고" if transaction_type == "in" else "출고"} 처리가 완료되었습니다.',
            'new_stock': new_stock
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@spare_parts_bp.route('/spare-parts/stock-transaction', methods=['POST'])
@jwt_required()
def create_stock_transaction():
    """재고 입출고 처리"""
    try:
        # 현재 사용자 정보 가져오기
        current_user_id = get_jwt_identity()
        current_user = User.get_by_id(int(current_user_id))
        user_name = current_user.name if current_user else 'Unknown'
        
        data = request.get_json()
        part_number = data.get('part_number')
        transaction_type = data.get('transaction_type')  # 'IN' 또는 'OUT'
        quantity = data.get('quantity', 0)
        transaction_date = data.get('transaction_date')
        reference_number = data.get('reference_number', '')
        customer_name = data.get('customer_name', '')
        
        if not all([part_number, transaction_type, quantity > 0]):
            return jsonify({
                'success': False,
                'error': '필수 필드가 누락되었습니다.'
            }), 400
        
        conn = get_db_connection()
        
        # 기존 파트 확인
        existing_part = conn.execute(
            'SELECT * FROM spare_parts WHERE part_number = ?', 
            (part_number,)
        ).fetchone()
        
        if not existing_part:
            conn.close()
            return jsonify({
                'success': False,
                'error': '부품을 찾을 수 없습니다.'
            }), 404
        
        previous_stock = existing_part['stock_quantity']
        
        if transaction_type == 'IN':
            new_stock = previous_stock + quantity
        elif transaction_type == 'OUT':
            if previous_stock < quantity:
                conn.close()
                return jsonify({
                    'success': False,
                    'error': f'재고가 부족합니다. (현재 재고: {previous_stock})'
                }), 400
            new_stock = previous_stock - quantity
        else:
            conn.close()
            return jsonify({
                'success': False,
                'error': '잘못된 거래 유형입니다.'
            }), 400
        
        # 재고 업데이트
        conn.execute(
            'UPDATE spare_parts SET stock_quantity = ? WHERE part_number = ?',
            (new_stock, part_number)
        )
        
        # 거래 내역 기록 (customer_name과 reference_number 포함)
        conn.execute(
            '''INSERT INTO stock_history 
               (part_number, transaction_type, quantity, previous_stock, new_stock, 
                transaction_date, reference_number, customer_name, created_at, created_by) 
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            (part_number, transaction_type, quantity, previous_stock, new_stock,
             transaction_date or datetime.now().date(), reference_number, customer_name,
             datetime.now(), user_name)
        )
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'{transaction_type.lower()} 처리가 완료되었습니다.',
            'data': {
                'part_number': part_number,
                'transaction_type': transaction_type,
                'quantity': quantity,
                'previous_stock': previous_stock,
                'new_stock': new_stock,
                'reference_number': reference_number,
                'customer_name': customer_name
            }
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@spare_parts_bp.route('/spare-parts/<int:part_id>/price-history', methods=['GET'])
def get_price_history(part_id):
    """파트 가격 변경 이력 조회"""
    try:
        conn = get_db_connection()
        
        # 가격 히스토리 조회 (최신순)
        price_history = conn.execute(
            '''SELECT * FROM price_history 
               WHERE spare_part_id = ? 
               ORDER BY effective_date DESC, created_at DESC''',
            (part_id,)
        ).fetchall()
        
        conn.close()
        
        # 결과 변환
        history_list = []
        for price in price_history:
            history_list.append({
                'id': price['id'],
                'price': price['price'],
                'effective_date': price['effective_date'],
                'notes': price['notes'],
                'created_at': price['created_at'],
                'created_by': price['created_by'],
                'currency': price['currency'] if price['currency'] else 'KRW',
                'part_type': price['part_type'] if price['part_type'] else 'repair',
                'billing_price': price['billing_price'] if price['billing_price'] else 0
            })
            
        return jsonify({
            'success': True,
            'data': history_list
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@spare_parts_bp.route('/spare-parts/<int:part_id>/price-history', methods=['POST'])
@jwt_required()
def add_price_history(part_id):
    """파트 가격 추가 (다중 통화 지원)"""
    try:
        data = request.get_json()
        
        # 필수 필드 검증
        required_fields = ['price', 'effective_date', 'currency', 'part_type']
        for field in required_fields:
            if field not in data:
                return jsonify({
                    'success': False,
                    'error': f'{field} 필드가 필요합니다.'
                }), 400
        
        price = data['price']
        effective_date = data['effective_date']
        currency = data['currency']
        part_type = data['part_type']
        notes = data.get('notes', '')
        
        # 유효성 검증
        if price <= 0:
            return jsonify({
                'success': False,
                'error': '가격은 0보다 커야 합니다.'
            }), 400
        
        if currency not in ['KRW', 'EUR', 'USD']:
            return jsonify({
                'success': False,
                'error': '지원하지 않는 통화입니다. (KRW, EUR, USD만 지원)'
            }), 400

        if part_type not in ['repair', 'consumable']:
            return jsonify({
                'success': False,
                'error': '지원하지 않는 부품 타입입니다. (repair, consumable만 지원)'
            }), 400
        
        # 현재 사용자 정보 가져오기
        current_user_id = get_jwt_identity()
        conn = get_db_connection()
        
        # 사용자 이름 가져오기
        user = conn.execute('SELECT name FROM users WHERE id = ?', (current_user_id,)).fetchone()
        user_name = user['name'] if user else 'Unknown'
        
        # 파트 존재 확인
        part = conn.execute('SELECT * FROM spare_parts WHERE id = ?', (part_id,)).fetchone()
        if not part:
            conn.close()
            return jsonify({
                'success': False,
                'error': '해당 파트를 찾을 수 없습니다.'
            }), 404
        
        # 환율 정보 가져오기 (KRW가 아닌 경우)
        krw_cost_price = price  # 원화 기준 원가
        exchange_rate = 1.0
        
        if currency != 'KRW':
            exchange_rate_info = conn.execute(
                '''SELECT rate FROM exchange_rates 
                   WHERE currency_from = ? AND currency_to = "KRW" AND is_active = 1''',
                (currency,)
            ).fetchone()
            
            if not exchange_rate_info:
                conn.close()
                return jsonify({
                    'success': False,
                    'error': f'{currency} 통화의 환율 정보를 찾을 수 없습니다.'
                }), 400
            
            exchange_rate = exchange_rate_info['rate']
            krw_cost_price = price * exchange_rate
        
        # 2차 함수 팩터 정보 가져오기 (관리자 설정에서)
        factor_info = conn.execute(
            '''SELECT factor_a, factor_b, factor_c, min_price, max_price, min_factor, max_factor 
               FROM pricing_factors 
               WHERE part_type = ? AND currency = "KRW"''',
            (part_type,)
        ).fetchone()
        
        if not factor_info:
            # 기본값 설정 (혹시 설정이 없는 경우)
            if part_type == 'repair':
                factor_info = {
                    'factor_a': 0.0000001,
                    'factor_b': -0.000615608,
                    'factor_c': 2.149275123,
                    'min_price': 100,
                    'max_price': 3000,
                    'min_factor': 1.20,
                    'max_factor': 2.10
                }
            else:  # consumable
                factor_info = {
                    'factor_a': 0.0000001,
                    'factor_b': -0.0003,
                    'factor_c': 1.6,
                    'min_price': 5,
                    'max_price': 300,
                    'min_factor': 1.20,
                    'max_factor': 1.55
                }
        
        # 2차 함수 팩터 적용: final_price = a*x^2 + b*x + c
        a = factor_info['factor_a']
        b = factor_info['factor_b'] 
        c = factor_info['factor_c']
        min_price = factor_info['min_price']
        max_price = factor_info['max_price']
        min_factor = factor_info['min_factor'] if factor_info['min_factor'] else 1.20
        max_factor = factor_info['max_factor'] if factor_info['max_factor'] else (2.10 if part_type == 'repair' else 1.55)
        
        # 통화별 팩터 계산 로직
        if currency == 'KRW':
            # KRW 원가는 최소/최대가격 상관없이 마진율만 적용
            # 저장된 마진율 가져오기
            margin_setting = conn.execute(
                '''SELECT setting_value FROM spare_part_settings 
                   WHERE setting_key = "margin_rate"'''
            ).fetchone()
            
            if margin_setting and margin_setting['setting_value']:
                margin_rate = 1 + (int(margin_setting['setting_value']) / 100)  # 25% -> 1.25
                print(f"KRW 마진율 적용: {margin_setting['setting_value']}% -> {margin_rate}")
            else:
                margin_rate = 1.20  # 기본 마진율 20% (1 + 0.20)
                print(f"기본 마진율 적용: 20% -> {margin_rate}")
            
            final_price_krw = krw_cost_price * margin_rate
        else:
            # EUR/USD 원가는 EUR/USD 기준으로 최소/최대가격 비교해서 팩터 계산
            original_currency_price = price  # EUR 또는 USD 원가
            
            # EUR/USD 기준 최소/최대가격과 비교
            if original_currency_price < min_price:
                # 최소 가격 미만일 때 최대 팩터 적용
                final_price_krw = krw_cost_price * max_factor
            elif original_currency_price > max_price:
                # 최대 가격 초과일 때 최소 팩터 적용
                final_price_krw = krw_cost_price * min_factor
            else:
                # 정상 범위일 때 2차 함수 적용 (EUR/USD 가격 기준)
                factor = a * (original_currency_price ** 2) + b * original_currency_price + c
                final_price_krw = krw_cost_price * factor
        
        # 최종 가격이 음수가 되지 않도록 보정
        final_price_krw = max(krw_cost_price, final_price_krw)
        
        # 100원 단위에서 올림 처리
        import math
        rounded_billing_price = math.ceil(final_price_krw / 100) * 100
        
        # 가격 히스토리 추가 (원가와 계산된 청구가 함께 저장)
        conn.execute(
            '''INSERT INTO price_history 
               (spare_part_id, price, effective_date, notes, created_at, created_by, currency, part_type, billing_price) 
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            (part_id, price, effective_date, notes, datetime.now(), user_name, currency, part_type, rounded_billing_price)
        )
        
        # 스페어 파트의 현재 가격 업데이트 (계산된 청구가격으로)
        conn.execute(
            '''UPDATE spare_parts 
               SET price = ? 
               WHERE id = ? AND ? >= (
                   SELECT COALESCE(MAX(effective_date), '1900-01-01') 
                   FROM price_history 
                   WHERE spare_part_id = ?
               )''',
            (int(rounded_billing_price), part_id, effective_date, part_id)
        )
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': '가격이 성공적으로 추가되었습니다.',
            'data': {
                'spare_part_id': part_id,
                'input_cost_price': price,  # 입력한 원가
                'currency': currency,
                'part_type': part_type,
                'krw_cost_price': int(krw_cost_price),  # 원화 환산 원가
                'final_billing_price': int(rounded_billing_price),  # 100원 단위 올림 처리된 최종 청구가격
                'exchange_rate': exchange_rate,
                'quadratic_factors': {
                    'a': a,
                    'b': b, 
                    'c': c
                },
                'effective_date': effective_date,
                'notes': notes,
                'created_by': user_name
            }
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@spare_parts_bp.route('/spare-parts/service-search', methods=['POST'])
def search_part_for_service_report():
    """부품번호 또는 부품명으로 부품 검색 (서비스 리포트용 - 부분 일치)"""
    try:
        data = request.get_json()
        search_term = data.get('part_number', '').strip() if data else ''

        if not search_term:
            return jsonify({
                'success': True,
                'spare_parts': []
            }), 200

        conn = get_db_connection()

        # 부품번호 또는 부품명으로 부분 일치 검색 (최대 10개)
        parts = conn.execute('''
            SELECT * FROM spare_parts
            WHERE part_number LIKE ? OR part_name LIKE ?
            LIMIT 10
        ''', (f'%{search_term}%', f'%{search_term}%')).fetchall()

        spare_parts = []
        for part in parts:
            part_dict = dict(part)

            # 최신 청구가 조회
            latest_billing_price = conn.execute('''
                SELECT billing_price
                FROM price_history
                WHERE spare_part_id = ?
                ORDER BY effective_date DESC, created_at DESC
                LIMIT 1
            ''', (part_dict['id'],)).fetchone()

            charge_price = latest_billing_price['billing_price'] if latest_billing_price else 0

            spare_parts.append({
                'id': part_dict['id'],
                'part_number': part_dict['part_number'],
                'part_name': part_dict['part_name'],
                'charge_price': charge_price,
                'stock_quantity': part_dict['stock_quantity']
            })

        conn.close()

        return jsonify({
            'success': True,
            'spare_parts': spare_parts
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@spare_parts_bp.route('/spare-parts/process-service-parts', methods=['POST'])
def process_service_parts():
    """서비스 리포트 저장 시 부품 처리 (출고 및 신규 등록)"""
    try:
        # 임시로 JWT 인증 제거하고 기본 사용자 사용
        # current_user_id = get_jwt_identity()
        # user = User.get_by_id(int(current_user_id))
        # if not user:
        #     return jsonify({'error': '사용자를 찾을 수 없습니다.'}), 401
        
        # 임시 사용자 정보
        class TempUser:
            username = 'system'
        user = TempUser()
            
        data = request.get_json()
        print(f"[DEBUG] 부품 처리 요청 데이터: {data}")
        
        service_report_id = data.get('service_report_id')
        customer_name = data.get('customer_name', '')
        used_parts = data.get('used_parts', [])
        
        print(f"[DEBUG] service_report_id: {service_report_id}")
        print(f"[DEBUG] customer_name: {customer_name}")
        print(f"[DEBUG] used_parts count: {len(used_parts)}")
        
        if not service_report_id:
            print("[DEBUG] 서비스 리포트 ID 누락")
            return jsonify({
                'success': False,
                'error': '서비스 리포트 ID가 필요합니다.'
            }), 400
        
        conn = get_db_connection()

        # 이미 처리된 부품이 있는지 확인 (중복 출고 방지)
        existing_history = conn.execute('''
            SELECT COUNT(*) as count FROM stock_history
            WHERE notes LIKE ?
        ''', (f'%서비스 리포트 ID: {service_report_id}%',)).fetchone()

        if existing_history and existing_history['count'] > 0:
            conn.close()
            return jsonify({
                'success': True,
                'message': '이미 처리된 서비스 리포트입니다. 중복 출고를 방지했습니다.',
                'processed_parts': [],
                'already_processed': True
            }), 200

        # 서비스 리포트 정보 조회 (작성일과 작성자 정보 가져오기)
        service_report = conn.execute('''
            SELECT sr.service_date, sr.technician_id, u.name
            FROM service_reports sr
            LEFT JOIN users u ON sr.technician_id = u.id
            WHERE sr.id = ?
        ''', (service_report_id,)).fetchone()

        if not service_report:
            conn.close()
            return jsonify({
                'success': False,
                'error': '해당 서비스 리포트를 찾을 수 없습니다.'
            }), 404
            
        # 서비스 리포트의 작성일을 출고일로 사용
        service_date = service_report['service_date']
        technician_name = service_report['name'] or 'system'
        print(f"[DEBUG] 서비스 리포트 작성일: {service_date}")
        print(f"[DEBUG] 기술자(작성자): {technician_name}")
        
        # 날짜 형식 확인 및 변환 (현재 시간을 포함한 완전한 datetime으로)
        if isinstance(service_date, str) and len(service_date) == 10:
            # YYYY-MM-DD 형식인 경우 현재 시간을 추가
            current_time = datetime.now()
            formatted_datetime = f"{service_date} {current_time.strftime('%H:%M:%S')}"
        else:
            # 다른 형식인 경우 현재 날짜와 시간 사용
            formatted_datetime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        print(f"[DEBUG] 포맷된 날짜/시간: {formatted_datetime}")
        
        processed_parts = []
        
        for part_data in used_parts:
            part_number = part_data.get('part_number', '').strip()
            part_name = part_data.get('part_name', '').strip()
            quantity = int(part_data.get('quantity', 0))
            unit_price = float(part_data.get('unit_price', 0))
            
            if quantity <= 0:
                continue
                
            if part_number:
                # 파트번호가 있는 경우 기존 부품 검색
                existing_part = conn.execute('''
                    SELECT * FROM spare_parts WHERE part_number = ?
                ''', (part_number,)).fetchone()
                
                if existing_part:
                    # 기존 부품 출고 처리
                    spare_part_id = existing_part['id']
                    
                    # 재고 확인 (경고만, 마이너스 재고 허용)
                    current_stock = existing_part['stock_quantity']
                    if current_stock < quantity:
                        # 마이너스 재고 허용하되 기록
                        pass
                    
                    # 재고 업데이트
                    new_stock = current_stock - quantity
                    conn.execute('''
                        UPDATE spare_parts 
                        SET stock_quantity = ?, updated_at = ? 
                        WHERE id = ?
                    ''', (new_stock, datetime.now().isoformat(), spare_part_id))
                    
                    # 출고 내역 기록
                    conn.execute('''
                        INSERT INTO stock_history 
                        (part_number, transaction_type, quantity, previous_stock, new_stock,
                         transaction_date, customer_name, reference_number, notes, created_by)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        part_number,
                        'out',  # 출고
                        quantity,
                        current_stock,
                        new_stock,
                        formatted_datetime,  # 포맷된 날짜/시간 사용
                        customer_name,  # 사용처를 고객사명으로
                        customer_name,  # reference_number도 고객사명으로
                        f'서비스 리포트 ID: {service_report_id}',
                        technician_name  # 레포트 작성자(기술자)를 출고 요청자로
                    ))
                    
                    processed_parts.append({
                        'part_number': part_number,
                        'part_name': existing_part['part_name'],
                        'action': 'outbound',
                        'quantity': quantity,
                        'new_stock': new_stock
                    })
                    
                else:
                    # 신규 부품 등록
                    if not part_name:
                        return jsonify({
                            'success': False,
                            'error': f'파트번호 {part_number}의 파트명이 필요합니다.'
                        }), 400
                    
                    # 신규 부품 등록
                    cursor = conn.execute('''
                        INSERT INTO spare_parts 
                        (part_number, part_name, stock_quantity, created_at, updated_at)
                        VALUES (?, ?, ?, ?, ?)
                    ''', (
                        part_number,
                        part_name,
                        -quantity,  # 초기 재고를 마이너스로 설정 (출고부터 시작)
                        datetime.now().isoformat(),
                        datetime.now().isoformat()
                    ))
                    
                    spare_part_id = cursor.lastrowid
                    
                    # 출고 내역 기록
                    conn.execute('''
                        INSERT INTO stock_history 
                        (part_number, transaction_type, quantity, previous_stock, new_stock,
                         transaction_date, customer_name, reference_number, notes, created_by)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        part_number,
                        'out',  # 출고
                        quantity,
                        0,  # 신규 부품이므로 이전 재고는 0
                        -quantity,  # 새로운 재고 (마이너스)
                        formatted_datetime,  # 포맷된 날짜/시간 사용
                        customer_name,  # 사용처를 고객사명으로
                        customer_name,  # reference_number도 고객사명으로
                        f'서비스 리포트 ID: {service_report_id} (신규 부품 등록)',
                        technician_name  # 레포트 작성자(기술자)를 출고 요청자로
                    ))
                    
                    processed_parts.append({
                        'part_number': part_number,
                        'part_name': part_name,
                        'action': 'new_and_outbound',
                        'quantity': quantity,
                        'new_stock': -quantity
                    })
            else:
                # 파트번호 없이 파트명만 있는 경우 (임시 처리, 별도 기록)
                # 실제로는 파트번호가 있어야 하지만 예외적으로 허용
                if part_name:
                    processed_parts.append({
                        'part_number': '',
                        'part_name': part_name,
                        'action': 'manual_entry',
                        'quantity': quantity,
                        'note': '파트번호 없이 수동 입력된 부품'
                    })
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'{len(processed_parts)}개 부품이 처리되었습니다.',
            'processed_parts': processed_parts
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@spare_parts_bp.route('/spare-parts/process-invoice-parts', methods=['POST'])
@jwt_required()
def process_invoice_parts():
    """거래명세서 저장 시 부품 처리 (출고 및 신규 등록)"""
    try:
        # 현재 로그인한 사용자 정보 가져오기
        current_user_id = get_jwt_identity()
        from app.models.user import User
        user = User.get_by_id(int(current_user_id))
        
        if not user:
            return jsonify({
                'success': False,
                'error': '사용자를 찾을 수 없습니다.'
            }), 401
            
        data = request.get_json()
        print(f"[DEBUG] 부품 처리 요청 데이터: {data}")
        print(f"[DEBUG] 요청 사용자: {user.name} (ID: {user.id})")
        
        invoice_id = data.get('invoice_id')
        customer_name = data.get('customer_name', '')
        used_parts = data.get('used_parts', [])
        
        print(f"[DEBUG] invoice_id: {invoice_id}")
        print(f"[DEBUG] customer_name: {customer_name}")
        print(f"[DEBUG] used_parts count: {len(used_parts)}")
        
        if not invoice_id:
            print("[DEBUG] 거래명세서 ID 누락")
            return jsonify({
                'success': False,
                'error': '거래명세서 ID가 필요합니다.'
            }), 400
        
        conn = get_db_connection()

        # 이미 처리된 부품이 있는지 확인 (중복 출고 방지)
        existing_history = conn.execute('''
            SELECT COUNT(*) as count FROM stock_history
            WHERE notes LIKE ?
        ''', (f'%거래명세서 ID: {invoice_id}%',)).fetchone()

        if existing_history and existing_history['count'] > 0:
            conn.close()
            return jsonify({
                'success': True,
                'message': '이미 처리된 거래명세서입니다. 중복 출고를 방지했습니다.',
                'processed_parts': [],
                'already_processed': True
            }), 200

        # 거래명세서 정보 조회 (발행일 정보 가져오기)
        invoice = conn.execute('''
            SELECT issue_date
            FROM invoices
            WHERE id = ?
        ''', (invoice_id,)).fetchone()

        if not invoice:
            conn.close()
            return jsonify({
                'success': False,
                'error': '해당 거래명세서를 찾을 수 없습니다.'
            }), 404
            
        # 거래명세서의 발행일을 출고일로 사용
        issue_date = invoice['issue_date']
        created_by = user.name  # 현재 로그인한 사용자 이름 사용
        print(f"[DEBUG] 거래명세서 발행일: {issue_date}")
        print(f"[DEBUG] 출고 요청자: {created_by}")
        
        # 날짜 형식 확인 및 변환 (현재 시간을 포함한 완전한 datetime으로)
        if isinstance(issue_date, str) and len(issue_date) == 10:
            # YYYY-MM-DD 형식인 경우 현재 시간을 추가
            current_time = datetime.now()
            formatted_datetime = f"{issue_date} {current_time.strftime('%H:%M:%S')}"
        else:
            # 다른 형식인 경우 현재 날짜와 시간 사용
            formatted_datetime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        print(f"[DEBUG] 포맷된 날짜/시간: {formatted_datetime}")
        
        processed_parts = []
        
        for part_data in used_parts:
            part_number = part_data.get('part_number', '').strip()
            part_name = part_data.get('part_name', '').strip()
            quantity = int(part_data.get('quantity', 0))
            unit_price = float(part_data.get('unit_price', 0))
            
            if quantity <= 0:
                continue
                
            if part_number:
                # 파트번호가 있는 경우 기존 부품 검색
                existing_part = conn.execute('''
                    SELECT * FROM spare_parts WHERE part_number = ?
                ''', (part_number,)).fetchone()
                
                if existing_part:
                    # 기존 부품 출고 처리
                    spare_part_id = existing_part['id']
                    
                    # 재고 확인 (경고만, 마이너스 재고 허용)
                    current_stock = existing_part['stock_quantity']
                    if current_stock < quantity:
                        # 마이너스 재고 허용하되 기록
                        pass
                    
                    # 재고 업데이트
                    new_stock = current_stock - quantity
                    conn.execute('''
                        UPDATE spare_parts 
                        SET stock_quantity = ?, updated_at = ? 
                        WHERE id = ?
                    ''', (new_stock, datetime.now().isoformat(), spare_part_id))
                    
                    # 출고 내역 기록
                    conn.execute('''
                        INSERT INTO stock_history 
                        (part_number, transaction_type, quantity, previous_stock, new_stock,
                         transaction_date, customer_name, reference_number, notes, created_by)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        part_number,
                        'out',  # 출고
                        quantity,
                        current_stock,
                        new_stock,
                        formatted_datetime,  # 포맷된 날짜/시간 사용
                        customer_name,  # 사용처를 고객사명으로
                        f'Invoice-{invoice_id}',  # reference_number는 Invoice-ID 형식으로
                        f'거래명세서 ID: {invoice_id}',
                        created_by  # 명세서 작성자를 출고 요청자로
                    ))
                    
                    processed_parts.append({
                        'part_number': part_number,
                        'part_name': existing_part['part_name'],
                        'action': 'outbound',
                        'quantity': quantity,
                        'new_stock': new_stock
                    })
                    
                else:
                    # 신규 부품 등록
                    if not part_name:
                        return jsonify({
                            'success': False,
                            'error': f'파트번호 {part_number}의 파트명이 필요합니다.'
                        }), 400
                    
                    # 신규 부품 등록
                    cursor = conn.execute('''
                        INSERT INTO spare_parts 
                        (part_number, part_name, stock_quantity, created_at, updated_at)
                        VALUES (?, ?, ?, ?, ?)
                    ''', (
                        part_number,
                        part_name,
                        -quantity,  # 초기 재고를 마이너스로 설정 (출고부터 시작)
                        datetime.now().isoformat(),
                        datetime.now().isoformat()
                    ))
                    
                    spare_part_id = cursor.lastrowid
                    
                    # 출고 내역 기록
                    conn.execute('''
                        INSERT INTO stock_history 
                        (part_number, transaction_type, quantity, previous_stock, new_stock,
                         transaction_date, customer_name, reference_number, notes, created_by)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        part_number,
                        'out',  # 출고
                        quantity,
                        0,  # 신규 부품이므로 이전 재고는 0
                        -quantity,  # 새로운 재고 (마이너스)
                        formatted_datetime,  # 포맷된 날짜/시간 사용
                        customer_name,  # 사용처를 고객사명으로
                        f'Invoice-{invoice_id}',  # reference_number는 Invoice-ID 형식으로
                        f'거래명세서 ID: {invoice_id} (신규 부품 등록)',
                        created_by  # 명세서 작성자를 출고 요청자로
                    ))
                    
                    processed_parts.append({
                        'part_number': part_number,
                        'part_name': part_name,
                        'action': 'new_and_outbound',
                        'quantity': quantity,
                        'new_stock': -quantity
                    })
            else:
                # 파트번호 없이 파트명만 있는 경우 (임시 처리, 별도 기록)
                if part_name:
                    processed_parts.append({
                        'part_number': '',
                        'part_name': part_name,
                        'action': 'manual_entry',
                        'quantity': quantity,
                        'note': '파트번호 없이 수동 입력된 부품'
                    })
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'{len(processed_parts)}개 부품이 처리되었습니다.',
            'processed_parts': processed_parts
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@spare_parts_bp.route('/spare-parts/inventory/monthly-summary', methods=['GET'])
@jwt_required()
def get_monthly_inventory_summary():
    """연도별 월별 재고 입출고 현황 조회"""
    try:
        year = request.args.get('year', date.today().year, type=int)

        conn = get_db_connection()
        cursor = conn.cursor()

        # 모든 파트 조회
        cursor.execute("""
            SELECT part_number, part_name, erp_name
            FROM spare_parts
            ORDER BY part_number
        """)
        parts = cursor.fetchall()

        result = []
        for part in parts:
            # 이월재고 계산 (선택 연도 이전 전체 입고 - 출고)
            cursor.execute("""
                SELECT
                    COALESCE(SUM(CASE WHEN transaction_type = 'IN' THEN quantity ELSE 0 END), 0) -
                    COALESCE(SUM(CASE WHEN transaction_type = 'OUT' THEN quantity ELSE 0 END), 0) as stock
                FROM stock_history
                WHERE part_number = ?
                  AND strftime('%Y', transaction_date) < ?
            """, (part['part_number'], str(year)))
            previous_year_stock = int(cursor.fetchone()['stock'])

            # 현재 재고 계산 (전체 입고 - 출고)
            cursor.execute("""
                SELECT
                    COALESCE(SUM(CASE WHEN transaction_type = 'IN' THEN quantity ELSE 0 END), 0) -
                    COALESCE(SUM(CASE WHEN transaction_type = 'OUT' THEN quantity ELSE 0 END), 0) as stock
                FROM stock_history
                WHERE part_number = ?
            """, (part['part_number'],))
            current_stock = int(cursor.fetchone()['stock'])

            # 이월재고 또는 현재재고가 0보다 큰 경우만 포함
            if previous_year_stock > 0 or current_stock > 0:
                monthly_data = {}

                # 각 월별 입출고 데이터 조회
                for month in range(1, 13):
                    # 입고 수량
                    cursor.execute("""
                        SELECT COALESCE(SUM(quantity), 0) as total
                        FROM stock_history
                        WHERE part_number = ?
                          AND transaction_type = 'IN'
                          AND strftime('%Y', transaction_date) = ?
                          AND strftime('%m', transaction_date) = ?
                    """, (part['part_number'], str(year), f'{month:02d}'))
                    inbound = cursor.fetchone()['total']

                    # 출고 수량
                    cursor.execute("""
                        SELECT COALESCE(SUM(quantity), 0) as total
                        FROM stock_history
                        WHERE part_number = ?
                          AND transaction_type = 'OUT'
                          AND strftime('%Y', transaction_date) = ?
                          AND strftime('%m', transaction_date) = ?
                    """, (part['part_number'], str(year), f'{month:02d}'))
                    outbound = cursor.fetchone()['total']

                    monthly_data[str(month)] = {
                        'inbound': int(inbound),
                        'outbound': int(outbound)
                    }

                result.append({
                    'part_number': part['part_number'],
                    'part_name': part['part_name'],
                    'erp_name': part['erp_name'],
                    'previous_year_stock': previous_year_stock,
                    'current_stock': current_stock,
                    'monthly_data': monthly_data
                })

        conn.close()

        return jsonify({
            'success': True,
            'data': {
                'year': year,
                'parts': result
            }
        }), 200

    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

@spare_parts_bp.route('/spare-parts/inventory/monthly-summary/export', methods=['GET'])
@jwt_required()
def export_monthly_inventory_summary():
    """월별 재고 현황 엑셀 파일 생성 및 다운로드"""
    try:
        year = request.args.get('year', date.today().year, type=int)

        # 인스턴스/연도별재고현황 폴더 생성
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
        export_dir = os.path.join(base_dir, 'instance', '연도별재고현황')
        os.makedirs(export_dir, exist_ok=True)

        # 파일명 생성
        filename = f'{year}-재고현황.xlsx'
        filepath = os.path.join(export_dir, filename)

        # 엑셀 워크북 생성
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f'{year}년 재고현황'

        # 스타일 정의
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')

        # 헤더 작성
        headers = ['파트번호', '파트명', 'ERP명', '이월재고', '현재재고']
        for month in range(1, 13):
            headers.extend([f'{month}월 입고', f'{month}월 출고'])

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border

        # 데이터 조회 및 작성
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT part_number, part_name, erp_name
            FROM spare_parts
            ORDER BY part_number
        """)
        parts = cursor.fetchall()

        row_num = 2
        for part in parts:
            # 이월재고 계산
            cursor.execute("""
                SELECT
                    COALESCE(SUM(CASE WHEN transaction_type = 'IN' THEN quantity ELSE 0 END), 0) -
                    COALESCE(SUM(CASE WHEN transaction_type = 'OUT' THEN quantity ELSE 0 END), 0) as stock
                FROM stock_history
                WHERE part_number = ?
                  AND strftime('%Y', transaction_date) < ?
            """, (part['part_number'], str(year)))
            previous_year_stock = int(cursor.fetchone()['stock'])

            # 현재 재고 계산
            cursor.execute("""
                SELECT
                    COALESCE(SUM(CASE WHEN transaction_type = 'IN' THEN quantity ELSE 0 END), 0) -
                    COALESCE(SUM(CASE WHEN transaction_type = 'OUT' THEN quantity ELSE 0 END), 0) as stock
                FROM stock_history
                WHERE part_number = ?
            """, (part['part_number'],))
            current_stock = int(cursor.fetchone()['stock'])

            # 이월재고 또는 현재재고가 0보다 큰 경우만 포함
            if previous_year_stock > 0 or current_stock > 0:
                # 기본 정보
                ws.cell(row=row_num, column=1, value=part['part_number']).border = border
                ws.cell(row=row_num, column=2, value=part['part_name']).border = border
                ws.cell(row=row_num, column=3, value=part['erp_name'] or '').border = border
                ws.cell(row=row_num, column=4, value=previous_year_stock).border = border
                ws.cell(row=row_num, column=5, value=current_stock).border = border

                # 월별 입출고 데이터
                col_num = 6
                for month in range(1, 13):
                    # 입고 수량
                    cursor.execute("""
                        SELECT COALESCE(SUM(quantity), 0) as total
                        FROM stock_history
                        WHERE part_number = ?
                          AND transaction_type = 'IN'
                          AND strftime('%Y', transaction_date) = ?
                          AND strftime('%m', transaction_date) = ?
                    """, (part['part_number'], str(year), f'{month:02d}'))
                    inbound = cursor.fetchone()['total']

                    # 출고 수량
                    cursor.execute("""
                        SELECT COALESCE(SUM(quantity), 0) as total
                        FROM stock_history
                        WHERE part_number = ?
                          AND transaction_type = 'OUT'
                          AND strftime('%Y', transaction_date) = ?
                          AND strftime('%m', transaction_date) = ?
                    """, (part['part_number'], str(year), f'{month:02d}'))
                    outbound = cursor.fetchone()['total']

                    in_cell = ws.cell(row=row_num, column=col_num, value=int(inbound))
                    in_cell.border = border
                    in_cell.alignment = center_align

                    out_cell = ws.cell(row=row_num, column=col_num + 1, value=int(outbound))
                    out_cell.border = border
                    out_cell.alignment = center_align

                    col_num += 2

                row_num += 1

        conn.close()

        # 열 너비 조정
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        for col in range(6, 6 + 24):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12

        # 파일 저장
        wb.save(filepath)

        return send_file(
            filepath,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500
